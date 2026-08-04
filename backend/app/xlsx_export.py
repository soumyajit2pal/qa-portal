"""Shared, audit-friendly Excel export helpers for test management."""

import datetime
import io
import re
from typing import Iterable, Mapping, Optional, Sequence

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "11182D"
PURPLE = "6557E8"
PALE_PURPLE = "ECE9FF"
PALE_BLUE = "EAF2FF"
PALE_GREEN = "DCFCE7"
PALE_RED = "FEE2E2"
PALE_AMBER = "FEF3C7"
TEXT = "172033"
MUTED = "667085"
WHITE = "FFFFFF"
GRID = "D8DEE9"
_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def safe_cell(value):
    """Make values Excel-safe while retaining dates and numeric values."""
    if isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, (datetime.date, int, float, bool)) or value is None:
        return value
    text = _ILLEGAL_XML.sub("", str(value))[:32767]
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def add_summary_sheet(
    workbook: Workbook,
    title: str,
    subtitle: str,
    metadata: Sequence[tuple[str, object]],
    metrics: Sequence[tuple[str, object]],
):
    ws = workbook.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = safe_cell(title)
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:F2")
    ws["A2"] = safe_cell(subtitle)
    ws["A2"].font = Font(size=10, color=MUTED)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28

    ws["A4"] = "Export details"
    ws["A4"].font = Font(size=12, bold=True, color=TEXT)
    row = 5
    for label, value in metadata:
        ws.cell(row, 1, safe_cell(label)).font = Font(bold=True, color=MUTED)
        ws.cell(row, 2, safe_cell(value)).alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(value, datetime.datetime):
            ws.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
        elif isinstance(value, datetime.date):
            ws.cell(row, 2).number_format = "yyyy-mm-dd"
        row += 1

    metric_row = max(row + 2, 13)
    ws.cell(metric_row, 1, "Snapshot metrics").font = Font(size=12, bold=True, color=TEXT)
    metric_row += 1
    for index, (label, value) in enumerate(metrics):
        col = (index % 3) * 2 + 1
        current_row = metric_row + (index // 3) * 3
        cell = ws.cell(current_row, col, safe_cell(label))
        cell.font = Font(size=9, bold=True, color=MUTED)
        value_cell = ws.cell(current_row + 1, col, safe_cell(value))
        value_cell.font = Font(size=18, bold=True, color=PURPLE)
        value_cell.fill = PatternFill("solid", fgColor=PALE_PURPLE)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=current_row + 1, start_column=col,
                       end_row=current_row + 1, end_column=col + 1)
        ws.cell(current_row + 1, col + 1).fill = PatternFill("solid", fgColor=PALE_PURPLE)
        ws.row_dimensions[current_row + 1].height = 28

    for col, width in {"A": 24, "B": 28, "C": 4, "D": 24, "E": 28, "F": 4}.items():
        ws.column_dimensions[col].width = width
    return ws


def add_table_sheet(
    workbook: Workbook,
    name: str,
    title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    *,
    subtitle: Optional[str] = None,
    widths: Optional[Mapping[str, float]] = None,
    wrap_headers: Optional[set[str]] = None,
    date_headers: Optional[set[str]] = None,
    date_only_headers: Optional[set[str]] = None,
    status_headers: Optional[set[str]] = None,
):
    rows = list(rows)
    ws = workbook.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    last_col = max(1, len(headers))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1, safe_cell(title))
    ws.cell(1, 1).font = Font(size=16, bold=True, color=WHITE)
    for column in range(1, last_col + 1):
        ws.cell(1, column).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1, safe_cell(subtitle or f"{len(rows)} record(s) in this sheet"))
    ws.cell(2, 1).font = Font(size=9, color=MUTED)

    header_row = 4
    thin = Side(style="thin", color=GRID)
    for column, header in enumerate(headers, 1):
        cell = ws.cell(header_row, column, safe_cell(header))
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
    ws.row_dimensions[header_row].height = 30

    wrap_headers = wrap_headers or set()
    date_headers = date_headers or set()
    date_only_headers = date_only_headers or set()
    status_headers = status_headers or set()
    status_fills = {
        "active": PALE_GREEN, "approved": PALE_GREEN, "pass": PALE_GREEN,
        "retest passed": PALE_GREEN, "fail": PALE_RED, "failed": PALE_RED,
        "blocked": PALE_AMBER, "draft": PALE_AMBER, "not executed": PALE_BLUE,
    }
    for row_number, values in enumerate(rows, header_row + 1):
        for column, header in enumerate(headers, 1):
            value = values[column - 1] if column - 1 < len(values) else None
            cell = ws.cell(row_number, column, safe_cell(value))
            cell.alignment = Alignment(
                vertical="top", wrap_text=header in wrap_headers,
            )
            cell.border = Border(bottom=thin)
            if row_number % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7F8FC")
            if header in date_headers and isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif header in date_only_headers and isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"
            if header in status_headers:
                fill = status_fills.get(str(cell.value or "").lower())
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
                    cell.font = Font(bold=True, color=TEXT)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{max(header_row, header_row + len(rows))}"
    widths = widths or {}
    for column, header in enumerate(headers, 1):
        sample_lengths = [len(str(safe_cell(header)))]
        sample_lengths.extend(
            len(str(safe_cell(row[column - 1] if column - 1 < len(row) else "") or ""))
            for row in rows[:250]
        )
        calculated = min(max(max(sample_lengths, default=10) + 2, 11), 44)
        ws.column_dimensions[get_column_letter(column)].width = widths.get(header, calculated)
    return ws


def workbook_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
