import io
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl

from .. import cache, models, schemas
from ..database import get_db
from ..deps import get_current_user, require_roles, require_same_department
from ..constants import Role, GatewayStatus
# _finalize_child_requests is the same child-creation step submit_request
# uses for the immediate (no approval needed) case -- reused here for the
# deferred case, once a brand-new Application Name clears this tier. No
# other router imports from another router anywhere else in this app; this
# is the one deliberate exception, since "once approved, then child request
# will be generated and will assign to SM" is squarely an Application Name
# decision triggering QA Request behaviour, not the other way around.
from .qa_requests import _finalize_child_requests

router = APIRouter(prefix="/api/application-names", tags=["application-names"])

# CAC-001..007 -- same rationale as departments.py: this is the QA Request
# wizard's Application Name dropdown, read very frequently and changed only
# when a name clears approval or an Admin bulk-seeds a batch.
_APPROVED_NAMES_CACHE_KEY = "refdata:application-names:approved:v1"
_APPROVED_NAMES_CACHE_TTL = 300


def _invalidate_approved_names_cache() -> None:
    cache.delete(_APPROVED_NAMES_CACHE_KEY)


def _log_application_name_decision(db: Session, obj: "models.ApplicationMaster", tier_label: str,
                                    decision: str, current_user: models.User, comments: str) -> None:
    """Reported directly: an Application Name Approve/Reject was never
    showing up on any request's own Activity tab -- only a REJECT happened
    to leave a trace, and only indirectly, via _auto_reject_linked_requests
    below force-rejecting whatever child request was still sitting at its
    own SM Approval (a "SM Approval / Rejected" entry, which reads as the
    request being rejected, not as the application name being rejected).
    Approvals left no trace anywhere. Fixed by logging this decision
    directly, against every request it's actually relevant to: every QA
    Request gateway that resolved to this exact ApplicationMaster row (a
    name can be reused as "Other" across more than one separately-raised QA
    Request, same set this function's callers already walk for the reject
    cascade), plus every one of that gateway's own linked Functional/SAST/
    DAST/Performance requests -- since the App Owner/SM banner and the
    pending/rejected badge are shown on every one of those screens (see
    ApplicationNameBanner and the Overview tab badges), so this decision
    should be visible in every one of those screens' own Activity tab too,
    not just the one screen the actual decision was made from."""
    step_name = f"Application Name ({tier_label})"
    for gw in db.query(models.QARequest).filter(models.QARequest.application_master_id == obj.id).all():
        db.add(models.ApprovalAction(
            entity_type="QA_REQUEST", entity_id=gw.id, step_name=step_name,
            actor_id=current_user.id, actor_role=current_user.roles_csv,
            decision=decision, comments=comments,
        ))
        # Query children directly instead of reading the ORM relationship
        # collections. During an Application Owner approval,
        # _finalize_child_requests() may have created these children moments
        # earlier after _sync_linked_child_requests() already loaded the
        # relationship collections as empty. Those cached empty collections
        # then made the activity fan-out silently skip every new child even
        # though the rows existed. Direct queries see both newly flushed and
        # pre-existing children consistently.
        groups = [
            (db.query(models.FunctionalRequest).filter_by(qa_request_id=gw.id).all(), "FUNCTIONAL_REQUEST"),
            (db.query(models.SASTRequest).filter_by(qa_request_id=gw.id).all(), "SAST"),
            (db.query(models.DASTRequest).filter_by(qa_request_id=gw.id).all(), "DAST"),
            (db.query(models.PerformanceRequest).filter_by(qa_request_id=gw.id).all(), "PERFORMANCE"),
        ]
        for children, entity_type in groups:
            for child in children:
                db.add(models.ApprovalAction(
                    entity_type=entity_type, entity_id=child.id, step_name=step_name,
                    actor_id=current_user.id, actor_role=current_user.roles_csv,
                    decision=decision, comments=comments,
                ))


def _auto_reject_linked_requests(db: Session, qa_request, current_user: models.User, comments: str):
    """Called when an Application Name is Rejected. Rejecting the name means
    the request(s) that introduced/use it can't legitimately proceed under
    it -- so every linked child request (Functional/SAST/DAST/
    Performance) still sitting at its own SM Approval checkpoint is force-
    rejected here too, rather than leaving the SM able to separately approve
    that request and send it on to Department Head with a name that was just
    rejected. Requests that have already moved past SM Approval (or haven't
    reached it yet) are left untouched -- this only closes the specific race
    the SM Approval screen otherwise allows (decide the name, then still
    freely decide the request itself)."""
    if not qa_request:
        return
    groups = [
        (qa_request.linked_functional_requests, "FUNCTIONAL_REQUEST"),
        (qa_request.linked_sast_requests, "SAST"),
        (qa_request.linked_dast_requests, "DAST"),
        (qa_request.linked_performance_requests, "PERFORMANCE"),
    ]
    for children, entity_type in groups:
        for child in children:
            if child.status == "SM_APPROVAL_PENDING":
                child.status = "SM_REJECTED"
                db.add(models.ApprovalAction(
                    entity_type=entity_type, entity_id=child.id, step_name="SM Approval",
                    actor_id=current_user.id, actor_role=current_user.roles_csv,
                    decision="Rejected", comments=comments,
                ))


def _approve_pending_application_name(db: Session, obj: "models.ApplicationMaster",
                                       current_user: models.User, comments: Optional[str]) -> None:
    """Shared "make this PENDING_APP_OWNER/PENDING_SM row terminal as
    Approved" state change -- used by decide_app_owner_name's own Approved
    branch (an Application Owner deciding one specific name) and by
    bulk_seed_application_names below (an Admin's Excel upload asserting a
    name is valid should also clear out that same name's own pending
    decision instead of leaving a stale duplicate queue entry sitting there
    once the name is already usable). Mirrors decide_app_owner_name's
    Approved branch exactly: terminal on both tiers' fields, and finalizes
    (creates linked children for) any QA Request gateway that introduced
    this name and is still sitting at Submitted, same as an Application
    Owner's own Approve would. Deliberately does NOT call
    _log_application_name_decision itself -- callers log it themselves,
    since decide_app_owner_name already does so unconditionally after this
    runs (for both Approve and Reject) and bulk-seeding wants its own
    distinct comment text on that activity entry."""
    now = models.now()
    obj.app_owner_decided_by_id = current_user.id
    obj.app_owner_decided_at = now
    obj.app_owner_comments = comments
    obj.status = "APPROVED"
    obj.decided_by_id = current_user.id
    obj.decided_at = now
    obj.comments = comments
    for gw in db.query(models.QARequest).filter(
            models.QARequest.application_master_id == obj.id,
            models.QARequest.status == GatewayStatus.SUBMITTED).all():
        _finalize_child_requests(db, gw, gw.requester)


@router.get("", response_model=List[schemas.ApplicationMasterOut])
def list_application_names(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """Approved names only -- this is what feeds the QA Request wizard's
    Application Name dropdown (see frontend QARequests/steps/DetailsStep.tsx).
    Anyone logged in can see the full approved list; there's nothing
    sensitive about a standardised application name."""
    cached = cache.get_json(_APPROVED_NAMES_CACHE_KEY)
    if cached is not None:
        return cached
    rows = (db.query(models.ApplicationMaster)
            .filter(models.ApplicationMaster.status == "APPROVED")
            .order_by(models.ApplicationMaster.name).all())
    result = [schemas.ApplicationMasterOut.model_validate(row).model_dump(mode="json") for row in rows]
    cache.set_json(_APPROVED_NAMES_CACHE_KEY, result, _APPROVED_NAMES_CACHE_TTL)
    return result


@router.get("/pending-app-owner", response_model=List[schemas.ApplicationMasterOut])
def list_pending_app_owner_names(db: Session = Depends(get_db),
                                  current_user: models.User = Depends(require_roles(Role.APPLICATION_OWNER, Role.ADMIN))):
    """Application Owner/Admin section listing every name still awaiting the
    FIRST tier of the two-tier approval chain -- mirrors
    list_pending_application_names below exactly, one tier earlier. An
    Application Owner's day-to-day path to this is normally the inline
    banner on their own module's detail view for the specific request that
    introduced the name (see application_master_status on FunctionalOut/
    SASTOut/DASTOut/PerformanceOut), not this list. ADMIN sees every
    department's pending names; an Application Owner only sees their own
    department's."""
    q = db.query(models.ApplicationMaster).filter(models.ApplicationMaster.status == "PENDING_APP_OWNER")
    if not current_user.has_role(Role.ADMIN):
        q = q.filter(models.ApplicationMaster.department == current_user.department)
    return q.order_by(models.ApplicationMaster.created_at).all()


@router.get("/pending", response_model=List[schemas.ApplicationMasterOut])
def list_pending_application_names(db: Session = Depends(get_db),
                                    current_user: models.User = Depends(require_roles(Role.SM, Role.ADMIN))):
    """SM/Admin section listing every name that has cleared Application Owner
    and is now awaiting SM, the second and final tier -- mainly useful for an
    Admin housekeeping view; an SM's day-to-day path to this is normally the
    inline banner on their own SM Approval screen for the specific request
    that introduced the name (see application_master_status on
    FunctionalOut/SASTOut/DASTOut/PerformanceOut), not this list. ADMIN sees
    every department's pending names; an SM only sees their own department's
    (same require_same_department scoping as the decision endpoint below,
    applied here as a filter instead of a hard error)."""
    q = db.query(models.ApplicationMaster).filter(models.ApplicationMaster.status == "PENDING_SM")
    if not current_user.has_role(Role.ADMIN):
        q = q.filter(models.ApplicationMaster.department == current_user.department)
    return q.order_by(models.ApplicationMaster.created_at).all()


@router.post("/{app_id}/app-owner-decision", response_model=schemas.ApplicationMasterOut)
def decide_app_owner_name(app_id: int, payload: schemas.ApplicationMasterDecision, db: Session = Depends(get_db),
                           current_user: models.User = Depends(require_roles(Role.APPLICATION_OWNER))):
    """Single-tier Application Name approval (2026-08 v2). Reported directly:
    "only application owner approval required, no SM involvement. if
    application owner approved then automatically come to SM for readiness
    verification and all" -- an Application Owner from the same department
    is now the ONLY decision this name ever needs; Approve is immediately
    terminal (moves straight to APPROVED, not PENDING_SM), and Reject is
    terminal too, same as before. This replaces the short-lived 2026-08
    two-tier chain (Application Owner, then a separate SM decision on the
    NAME itself) -- see models.ApplicationMaster's own docstring and
    decide_application_name below, which is now legacy-only (kept working
    for any pre-existing PENDING_SM row from before this change, but no new
    row can ever reach that status again; see the migration notes for the
    one-time data fix-up). Since Approve is now terminal, it also populates
    the SM-tier decided_by_id/decided_at/comments fields (not just this
    tier's own app_owner_* fields) -- same reasoning Reject already used
    ("the decision that made this terminal"), now true for both outcomes,
    not just Reject. Same same-department scoping as every other approval
    checkpoint in the app.

    2026-08: a brand-new name introduced on a QA Request gateway defers that
    gateway's own child-request creation until it clears THIS tier (see
    routers/qa_requests.py::submit_request) -- the gateway sits at Submitted,
    not yet Raised, with no linked Functional/SAST/DAST/Performance request
    of its own yet. So this decision has to drive that gateway forward too,
    not just the ApplicationMaster row itself: Approve finalizes every such
    gateway right here (children get created and assigned to SM for their
    own normal readiness verification -- "automatically come to SM for
    readiness verification and all", exactly as reported -- same as an
    immediate raise, see _finalize_child_requests), attributed to the
    ORIGINAL requester, not this Application Owner. Reject sends any such
    gateway all the way back to Draft instead -- since it never got as far as
    creating a single child, there's nothing for _auto_reject_linked_requests
    to do for it, and "awaiting approval forever with nothing to show for it"
    isn't a real state; the requester can simply edit and resubmit under a
    different name."""
    obj = db.query(models.ApplicationMaster).get(app_id)
    if not obj:
        raise HTTPException(404, "Application name not found")
    require_same_department(current_user, obj.department)
    if obj.status != "PENDING_APP_OWNER":
        raise HTTPException(
            400,
            f"This application name is not awaiting Application Owner decision -- "
            f"its current status is '{obj.status}'.",
        )
    if payload.decision not in ("Approved", "Rejected"):
        raise HTTPException(400, "decision must be one of: Approved, Rejected")
    if payload.decision == "Approved":
        # Terminal here now -- no more PENDING_SM tier for a NEW decision
        # (see docstring above). Mirrored into the SM-tier fields too, same
        # as Reject already did, so anything reading decided_by_id/
        # decided_at/comments as "the decision that made this terminal"
        # keeps working regardless of which outcome it was. Deferred
        # child-request creation (2026-08) for any gateway that introduced
        # this name and stopped at Submitted is handled inside the shared
        # helper -- see its own docstring.
        _approve_pending_application_name(db, obj, current_user, payload.comments)
    else:
        now = models.now()
        obj.app_owner_decided_by_id = current_user.id
        obj.app_owner_decided_at = now
        obj.app_owner_comments = payload.comments
        obj.status = "REJECTED"
        # Reject at this tier IS the final decision -- mirror it into the
        # SM-tier fields too, same reasoning as the docstring above.
        obj.decided_by_id = current_user.id
        obj.decided_at = now
        obj.comments = payload.comments
        reason = f"Application Name '{obj.name}' was rejected by Application Owner"
        if payload.comments:
            reason += f": {payload.comments}"
        for gw in db.query(models.QARequest).filter(models.QARequest.application_master_id == obj.id).all():
            if gw.status == GatewayStatus.SUBMITTED:
                # Never got as far as creating a single child request --
                # back to Draft outright rather than force-rejecting
                # requests that don't exist yet (there's nothing for
                # _auto_reject_linked_requests to do here). request_id
                # (already assigned at Submit) is deliberately KEPT, not
                # nulled out, for traceability -- a narrow, documented
                # exception to its column comment's usual "stays NULL while
                # Draft" rule on models.QARequest.
                gw.status = GatewayStatus.DRAFT
                db.add(models.ApprovalAction(
                    entity_type="QA_REQUEST", entity_id=gw.id, step_name="Requester",
                    actor_id=current_user.id, actor_role=current_user.roles_csv,
                    decision="Reverted to Draft",
                    comments=(
                        f"Application Name '{obj.name}' was rejected by Application Owner before any "
                        "linked request was generated -- edit and resubmit under a different name."
                    ),
                ))
            else:
                _auto_reject_linked_requests(db, gw, current_user, reason)
    _log_application_name_decision(db, obj, "Application Owner", payload.decision, current_user, payload.comments)
    db.commit()
    db.refresh(obj)
    _invalidate_approved_names_cache()
    return obj


@router.post("/{app_id}/decision", response_model=schemas.ApplicationMasterOut)
def decide_application_name(app_id: int, payload: schemas.ApplicationMasterDecision, db: Session = Depends(get_db),
                             current_user: models.User = Depends(require_roles(Role.SM))):
    """LEGACY-ONLY as of 2026-08 v2 (see decide_app_owner_name's own
    docstring -- reported directly: "only application owner approval
    required, no SM involvement"). Application Owner approval is now
    terminal on its own; no NEW ApplicationMaster row can ever reach
    PENDING_SM again, so this endpoint can only ever act on a row that was
    already sitting at PENDING_SM from before this change shipped (see the
    migration notes' one-time data fix-up, which converts any such row
    straight to APPROVED so this code path shouldn't normally be reachable
    at all post-migration; left in place, unremoved, purely as a safety net
    for any row the fix-up missed rather than leaving it permanently stuck).
    Approving never changes the linked request's own status -- an approved
    name simply becomes a standard dropdown option going forward. Rejecting
    is different: it also force-rejects any linked request still sitting at
    its own SM Approval checkpoint (see _auto_reject_linked_requests above)
    -- a request can't be allowed to proceed to Department Head under an
    application name the SM just rejected. Same same-department scoping as
    every other SM approval checkpoint in the app."""
    obj = db.query(models.ApplicationMaster).get(app_id)
    if not obj:
        raise HTTPException(404, "Application name not found")
    require_same_department(current_user, obj.department)
    if obj.status == "PENDING_APP_OWNER":
        raise HTTPException(
            400,
            "This application name is still awaiting Application Owner approval -- "
            "it hasn't reached SM review yet.",
        )
    if obj.status != "PENDING_SM":
        raise HTTPException(400, f"This application name has already been '{obj.status}'")
    if payload.decision not in ("Approved", "Rejected"):
        raise HTTPException(400, "decision must be one of: Approved, Rejected")
    obj.status = "APPROVED" if payload.decision == "Approved" else "REJECTED"
    obj.decided_by_id = current_user.id
    obj.decided_at = models.now()
    obj.comments = payload.comments
    if obj.status == "REJECTED":
        reason = f"Application Name '{obj.name}' was rejected by SM"
        if payload.comments:
            reason += f": {payload.comments}"
        # Walk every QA Request gateway that resolved to this exact
        # ApplicationMaster row -- not just obj.qa_request (the one gateway
        # that happened to introduce it first via _resolve_application_name).
        # A requester can reuse the same "Other" name across more than one
        # separately-raised QA Request; every one of them shares this same
        # application_master_id and needs its own linked requests rejected.
        for gw in db.query(models.QARequest).filter(models.QARequest.application_master_id == obj.id).all():
            _auto_reject_linked_requests(db, gw, current_user, reason)
    _log_application_name_decision(db, obj, "SM", payload.decision, current_user, payload.comments)
    db.commit()
    db.refresh(obj)
    _invalidate_approved_names_cache()
    return obj


# ---- Admin bulk-seed from Excel ----
# Reported directly: "add one functionality on admin section to upload excel
# and based on data present on excel Application name will be seed."

def _normalize_seed_header(h) -> str:
    return str(h or "").strip().lower()


_SEED_HEADER_MAP = {
    "application name": "name",
    "app name": "name",
    "name": "name",
    "department": "department",
}


@router.get("/bulk-seed-template")
def download_bulk_seed_template(current_user: models.User = Depends(require_roles(Role.ADMIN))):
    """Minimal xlsx template for bulk_seed_application_names below, built on
    the fly with openpyxl rather than a bundled static asset (contrast with
    test_repository.py's own /import-template, which serves a real static
    file since that template is a much larger fixed layout) -- this one is
    just a two-column header plus one example row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Application Names"
    ws.append(["Application Name", "Department"])
    ws.append(["EXAMPLE APPLICATION", "IT - Software"])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="application_names_seed_template.xlsx"'},
    )


@router.post("/bulk-seed", response_model=schemas.ApplicationSeedResult)
async def bulk_seed_application_names(file: UploadFile = File(...), db: Session = Depends(get_db),
                                       current_user: models.User = Depends(require_roles(Role.ADMIN))):
    """Admin-only bulk import: seeds ApplicationMaster directly from an xlsx
    of known-good application names.

    Unlike the normal path (a requester typing "Other" on the QA Request
    wizard -- see _resolve_application_name in qa_requests.py), a brand-new
    name seeded here is created straight at APPROVED. An Admin bulk-uploading
    a spreadsheet of already-known-valid application names is asserting
    they're valid, not proposing them for the usual Application Owner review
    -- there's nothing to route through that queue for a row with no
    originating QA Request at all (qa_request_id stays NULL, same as any
    other row with nothing to trace back to).

    An existing row still sitting at PENDING_APP_OWNER/PENDING_SM is instead
    approved outright (see _approve_pending_application_name, which also
    finalizes/creates linked children for any gateway request still waiting
    on this exact name) rather than creating a duplicate -- so re-running the
    same file after some of its names were separately proposed elsewhere
    clears those out too instead of erroring or duplicating. An existing
    APPROVED row is left untouched and counted as a duplicate. An existing
    REJECTED row is also left untouched (counted separately) rather than
    silently overridden -- a real Reject decision may have already
    force-rejected other linked requests (see _auto_reject_linked_requests),
    so reinstating one is left to the normal decision endpoints, not a bulk
    upload.

    Expects a header row with an "Application Name" column (case/whitespace
    tolerant, same convention as test_repository.py's own xlsx import) and an
    optional "Department" column; any other columns are ignored. See
    download_bulk_seed_template above for the expected shape."""
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(400, "Could not read this file as an Excel (.xlsx) workbook")
    ws = wb.worksheets[0]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(400, "Sheet has no header row")
    col_fields: dict = {}
    for idx, header in enumerate(header_row):
        field = _normalize_seed_header(header)
        field = _SEED_HEADER_MAP.get(field)
        if field and field not in col_fields.values():
            col_fields[idx] = field
    if "name" not in col_fields.values():
        raise HTTPException(400, "Could not find an 'Application Name' column -- is this the right template?")

    created = 0
    approved_existing = 0
    skipped_duplicate = 0
    skipped_rejected = 0
    skipped_invalid = 0
    errors: List[str] = []
    seen_in_file: set = set()
    seed_comment = f"Seeded via Admin bulk Excel upload by {current_user.full_name}"

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        parsed: dict = {}
        for idx, field in col_fields.items():
            if idx < len(row):
                value = row[idx]
                parsed[field] = str(value).strip() if value is not None else ""
        name = (parsed.get("name") or "").strip()
        if not name:
            skipped_invalid += 1
            errors.append(f"Row {row_number}: no Application Name value -- skipped.")
            continue
        name_upper = name.upper()
        department = (parsed.get("department") or "").strip() or None
        if name_upper in seen_in_file:
            skipped_duplicate += 1
            errors.append(f"Row {row_number}: '{name}' occurs more than once in this workbook -- later occurrence skipped.")
            continue
        seen_in_file.add(name_upper)

        existing = db.query(models.ApplicationMaster).filter(models.ApplicationMaster.name == name_upper).first()
        if existing:
            if existing.status == "APPROVED":
                skipped_duplicate += 1
            elif existing.status == "REJECTED":
                skipped_rejected += 1
                errors.append(
                    f"Row {row_number}: '{name}' was previously rejected and was left untouched -- "
                    "use the normal Application Name decision screen to reinstate it if that's intended."
                )
            else:
                _approve_pending_application_name(db, existing, current_user, seed_comment)
                _log_application_name_decision(db, existing, "Application Owner", "Approved", current_user, seed_comment)
                approved_existing += 1
            continue

        obj = models.ApplicationMaster(
            name=name_upper, status="APPROVED", department=department,
            requested_by_id=current_user.id, qa_request_id=None,
            app_owner_decided_by_id=current_user.id, app_owner_decided_at=models.now(),
            app_owner_comments=seed_comment,
            decided_by_id=current_user.id, decided_at=models.now(), comments=seed_comment,
        )
        db.add(obj)
        created += 1

    db.commit()
    if created or approved_existing:
        _invalidate_approved_names_cache()
    failure_reason = None
    if created == 0 and approved_existing == 0:
        failure_reason = errors[0] if errors else (
            "No application names were created or approved. Confirm the workbook has an 'Application Name' "
            "column with data below the header row."
        )
    return schemas.ApplicationSeedResult(
        created=created, approved_existing=approved_existing, skipped_duplicate=skipped_duplicate,
        skipped_rejected=skipped_rejected, skipped_invalid=skipped_invalid, errors=errors,
        failure_reason=failure_reason,
    )
