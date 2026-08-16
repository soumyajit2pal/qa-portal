import asyncio
import io
import os
from typing import List, Optional
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload, selectinload
import openpyxl

from .. import models, schemas, pagination
from ..database import get_db, SessionLocal
from ..deps import (
    get_current_user, require_roles, get_or_404,
    require_can_author_repository, require_can_review_repository, require_can_give_final_approval,
    require_can_manage_repository_governance,
    can_review_repository,
    get_project_or_404 as _get_project_or_404,
)
from ..constants import (
    Role, TEST_CASE_PRIORITIES, TEST_CASE_STATUSES,
    TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS,
)
from ..xlsx_export import add_summary_sheet, add_table_sheet, new_workbook, workbook_response
from . import jobs

# PAG-005 -- every eager-load the list endpoint needs to serialize
# TestCaseListOut without an N+1: folder/created_by/checked_out_by are
# simple one-to-one FKs; current_draft_version is itself joined one level
# further to whichever of assigned_reviewer/assigned_qa_lead/author its own
# pending_with_user_name property reads (see models.TestCaseVersion); tags
# are a real one-to-many table, selectinload'd (a second batched query, not
# a join that would multiply the case row count).
_LIST_CASE_EAGER_LOADS = [
    joinedload(models.TestCase.folder),
    joinedload(models.TestCase.created_by),
    joinedload(models.TestCase.checked_out_by),
    joinedload(models.TestCase.current_draft_version).joinedload(models.TestCaseVersion.assigned_reviewer),
    joinedload(models.TestCase.current_draft_version).joinedload(models.TestCaseVersion.assigned_qa_lead),
    joinedload(models.TestCase.current_draft_version).joinedload(models.TestCaseVersion.author),
    selectinload(models.TestCase.tag_rows),
]

# Canonical "Test Cases - CR-XX - Template" xlsx that import_test_cases()
# below parses -- shipped as a static, git-tracked asset (NOT the runtime
# uploads/ folder, which is excluded from deploy syncs) so the Test
# Repository "Download Template" button always has something to serve.
_IMPORT_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "assets", "templates", "Test_Case_Import_Template.xlsx",
)

router = APIRouter(prefix="/api/test-repository", tags=["test-management"])

# QA Engineer + QA Lead both author (create/edit/import/delete) test cases in
# the Repository, per direct product decision -- Admin always bypasses via
# require_roles. CHIEF_MANAGER_QA/AGM_QA also included (comment previously
# didn't mention them, though the tuple already did) -- both hold identical
# Author-tier standing here, same as Stage 2 approval (_stage2_approver_ids).
_AUTHOR_ROLES = (Role.QA_ENGINEER, Role.QA_LEAD, Role.CHIEF_MANAGER_QA, Role.AGM_QA)


def _stage1_reviewer_ids(db: Session, project_id: int, author_id: Optional[int] = None) -> List[int]:
    """OLD-path only (see TEST_CASE_NEW_STATUSES comment in constants.py) --
    all active members of the QA Lead group, excluding the author. Kept
    exactly as-is so drafts already sitting at "In Review" when the 2026-08
    "Simplified Test Management" change shipped keep working unchanged."""
    users = db.query(models.User).filter(
        # Oracle stores SQLAlchemy Boolean as NUMBER(1); `.is_(True)` emits
        # `IS 1`, which Oracle rejects with ORA-00908. Equality emits `= 1`.
        models.User.is_active == 1,
        # 2026-08 "one user can be on multiple departments" CR -- membership
        # via the department_assignments join table, not the legacy column.
        models.User.department_assignments.any(
            models.UserDepartment.department.in_(TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS)
        ),
    ).all()
    return [
        user.id for user in users
        if user.id != author_id and Role.QA_LEAD in set(user.roles)
    ]


def _stage2_approver_ids(db: Session, author_id: Optional[int] = None) -> List[int]:
    """OLD-path only -- all active CM QA and AGM QA users; either one may
    complete Stage 2. Kept exactly as-is for drafts already sitting at
    "Review Completed" -- see _stage1_reviewer_ids' own comment."""
    users = db.query(models.User).filter(
        models.User.is_active == 1,
        # 2026-08 "one user can be on multiple departments" CR -- membership
        # via the department_assignments join table, not the legacy column.
        models.User.department_assignments.any(
            models.UserDepartment.department.in_(TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS)
        ),
    ).all()
    return [
        user.id for user in users
        if user.id != author_id and set(user.roles).intersection({Role.CHIEF_MANAGER_QA, Role.AGM_QA})
    ]


def _qa_group_ids(db: Session, author_id: Optional[int] = None) -> List[int]:
    """NEW-path Stage 1 (Recommendation) -- "Any eligible QA member can
    recommend the test case" per the Simplified Test Management requirement,
    section 2's "QA Group" row. Mapped onto Role.QA_ENGINEER (confirmed via
    AskUserQuestion), excluding the author (GOV-002 maker-checker)."""
    users = db.query(models.User).filter(
        models.User.is_active == 1,
        # 2026-08 "one user can be on multiple departments" CR -- membership
        # via the department_assignments join table, not the legacy column.
        models.User.department_assignments.any(
            models.UserDepartment.department.in_(TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS)
        ),
    ).all()
    return [
        user.id for user in users
        if user.id != author_id and Role.QA_ENGINEER in set(user.roles)
    ]


def _qa_lead_group_ids(db: Session, author_id: Optional[int] = None) -> List[int]:
    """NEW-path Stage 2 (Final Approval) -- "Any eligible QA Lead can
    approve" per the Simplified Test Management requirement, section 2's "QA
    Lead Group" row. Mapped onto QA_LEAD/CHIEF_MANAGER_QA/AGM_QA (confirmed
    via AskUserQuestion -- the same Executive-bypass group used for QA-Lead-
    gated actions elsewhere, see ORACLE_MIGRATION_2026-07.md section 59),
    excluding the author."""
    users = db.query(models.User).filter(
        models.User.is_active == 1,
        # 2026-08 "one user can be on multiple departments" CR -- membership
        # via the department_assignments join table, not the legacy column.
        models.User.department_assignments.any(
            models.UserDepartment.department.in_(TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS)
        ),
    ).all()
    return [
        user.id for user in users
        if user.id != author_id
        and set(user.roles).intersection({Role.QA_LEAD, Role.CHIEF_MANAGER_QA, Role.AGM_QA})
    ]


# Drafts already sitting in one of these OLD-vocabulary statuses when the
# Simplified Test Management change shipped keep running the pre-existing,
# untouched individually-routed logic to completion -- "new cases only" per
# AskUserQuestion. Anything else (a fresh Draft submission, or a NEW-
# vocabulary Returned status resubmitting) uses the new QA Group/QA Lead
# Group logic. See TEST_CASE_NEW_STATUSES in constants.py.
_OLD_WORKFLOW_STATUSES = {"In Review", "Review Completed", "Returned"}


def _lock_current_drafts(db: Session, cases: List[models.TestCase]) -> None:
    """Serialize bulk workflow decisions so the first committed action wins."""
    draft_ids = sorted({case.current_draft_version_id for case in cases if case.current_draft_version_id})
    if draft_ids:
        db.query(models.TestCaseVersion).filter(
            models.TestCaseVersion.id.in_(draft_ids)
        ).order_by(models.TestCaseVersion.id).populate_existing().with_for_update().all()

# ---------------------------------------------------------------------------
# 2026-08 "Test Management Revamp" -- see models.py's own "Module 10" header
# comment for the full architecture. Every testcase content edit below now
# targets an immutable-once-approved TestCaseVersion row instead of the
# TestCase identity's own columns directly (VER-001/VER-003/VER-006); this
# router's job is to keep TestCase's own mirror columns
# (status/version_major/version_minor/content) in sync with whichever
# version is "current" for display/list-view backward compatibility -- see
# _sync_case_mirror below, called after every state change.
# ---------------------------------------------------------------------------

_CONTENT_FIELDS = [
    "epic_id", "cr_number", "feature_id", "user_story_id", "test_type", "module_name",
    "test_scenario", "pre_condition", "description", "priority",
]


def _case_workflow_action(case_id: int, current_user: models.User, decision: str,
                          comments: Optional[str] = None,
                          previous_state: Optional[str] = None, new_state: Optional[str] = None) -> models.ApprovalAction:
    """APR-005 -- previous_state/new_state are optional (see ApprovalAction's
    own docstring) but populated whenever the caller is recording an actual
    TestCaseVersion status transition, closing the literal state-machine
    traceability requirement for this entity type."""
    return models.ApprovalAction(
        entity_type="TEST_CASE", entity_id=case_id, step_name="Test Case Approval Workflow",
        actor_id=current_user.id, actor_role=current_user.roles_csv,
        decision=decision, comments=comments,
        previous_state=previous_state, new_state=new_state,
    )


def _require_active_project(project: models.TestProject) -> None:
    if not project.is_active:
        raise HTTPException(400, "This Test Project is inactive. Reactivate it before changing repository content")


def _sync_case_mirror(case: models.TestCase, version: models.TestCaseVersion) -> None:
    """Copies one TestCaseVersion's content/status/numbers onto its parent
    TestCase's own mirror columns -- called after every state change so
    every pre-revamp list/filter view that reads TestCase's own columns
    directly (rather than joining to its versions) keeps showing accurate
    data. "Current" is always whichever version this function was last
    called with -- callers decide precedence (draft-if-any, else approved),
    see each endpoint below."""
    case.status = version.status
    case.version_major = version.version_major
    case.version_minor = version.version_minor
    for field in _CONTENT_FIELDS:
        setattr(case, field, getattr(version, field))


def _current_display_version(case: models.TestCase) -> Optional[models.TestCaseVersion]:
    """Whichever version should currently be shown/mirrored: the in-progress
    draft if one exists (so the UI shows what's being worked on), else the
    approved baseline, else nothing (shouldn't happen once a case has been
    through creation, but defensive)."""
    return case.current_draft_version or case.current_approved_version


def _next_provisional_version_numbers(case: models.TestCase) -> tuple:
    """Provisional version_major/minor assigned when a new Draft is first
    created -- provisional because SRS VER-004 lets the QA Lead choose a
    major bump instead at APPROVAL time (see review_test_case), at which
    point these get recomputed. A case with no approved baseline yet used
    to always start at 1.0 (VER-004/original versioning comment: "a newly
    created case starts at 1.0... stays 1.0 until first approved") --
    ORA-00001 fix: that's only true the FIRST time. Reject is terminal but
    does NOT clear TestCase.current_draft_version_id (see review_test_case's
    REJECT branch), so a case whose only version so far was Rejected still
    has "no approved baseline" yet already has a real row sitting at (1, 0)
    in qap_test_case_versions -- editing it again (update_test_case's
    rejected_base path / bulk_update_test_cases) would otherwise try to
    INSERT a second row at that same (test_case_id, version_major,
    version_minor), violating UQ_QAP_TCV_CASE_VERSION. Look at the actual
    version history instead of inferring "nothing exists yet" from "nothing
    approved yet" -- those are only the same thing before a first Reject."""
    approved = case.current_approved_version
    if approved:
        return approved.version_major, approved.version_minor + 1
    existing = case.versions
    if not existing:
        return 1, 0
    highest = max(existing, key=lambda v: (v.version_major, v.version_minor))
    return highest.version_major, highest.version_minor + 1


def _validate_steps(steps: List[models.TestCaseVersionStep]) -> None:
    """TC-003 -- reject empty step text, duplicate step numbers, or a step
    with no expected result. Called at submit time, not at every draft
    save, so authors can save incomplete work-in-progress freely (VER-001:
    a version only truly matters once it's been submitted for review)."""
    if not steps:
        raise HTTPException(400, "Add at least one step before submitting for review")
    seen_numbers = set()
    for step in steps:
        if step.step_no in seen_numbers:
            raise HTTPException(400, f"Duplicate step number {step.step_no} -- each step must have a unique number")
        seen_numbers.add(step.step_no)
        if not (step.step_text or "").strip():
            raise HTTPException(400, f"Step {step.step_no} cannot have blank step text")
        if not (step.expected_result or "").strip():
            raise HTTPException(400, f"Step {step.step_no} is missing an expected result")


def _replace_draft_steps(db: Session, case: models.TestCase, version: models.TestCaseVersion,
                         steps: List[schemas.TestStepIn]) -> None:
    """Replace Draft steps without colliding with Oracle's unique key.

    A relationship assignment can batch replacement INSERTs before orphan
    DELETEs. Oracle then sees the old ``(version_id, step_no)`` row and
    raises ORA-00001. Flush an explicit delete phase before adding the new
    rows so the replacement is deterministic on every supported database.
    """
    seen_numbers = set()
    for step in steps:
        if step.step_no in seen_numbers:
            raise HTTPException(
                400,
                f"Duplicate step number {step.step_no} -- each step must have a unique number",
            )
        seen_numbers.add(step.step_no)

    version.steps.clear()
    # Maintain the legacy TestCase.steps compatibility mirror in the same
    # two-phase operation.
    case.steps.clear()
    db.flush()

    version.steps.extend(models.TestCaseVersionStep(**step.model_dump()) for step in steps)
    case.steps.extend(models.TestStep(**step.model_dump()) for step in steps)


def _create_case_with_first_draft(
    db: Session, project_id: int, folder_id: Optional[int], content: dict, tags: List[str],
    steps: List[dict], current_user: models.User, source_version_id: Optional[int] = None,
) -> models.TestCase:
    """Shared by create_test_case, import_test_cases, clone_test_case, and
    _clone_folder_subtree -- every path that mints a brand-new TestCase
    identity also needs its very first TestCaseVersion (1.0, Draft). Kept in
    one place so all four stay consistent instead of drifting."""
    key = models.gen_id(models.BUSINESS_ID_PREFIXES["TEST_CASE"], db)
    case = models.TestCase(project_id=project_id, folder_id=folder_id, test_case_key=key,
                            created_by_id=current_user.id, status="Draft", **content)
    case.tag_rows = [models.TestCaseTag(tag=tag) for tag in tags]
    db.add(case)
    db.flush()

    version = models.TestCaseVersion(
        test_case_id=case.id, version_major=1, version_minor=0, status="Draft",
        author_id=current_user.id, created_at=models.now(), source_version_id=source_version_id,
        **content,
    )
    db.add(version)
    db.flush()
    version.steps = [models.TestCaseVersionStep(step_no=s["step_no"], step_text=s.get("step_text"),
                                                 expected_result=s.get("expected_result")) for s in steps]
    # Kept on the legacy table too, purely so any code path that still reads
    # TestCase.steps (none should after this revamp, but see TestCase's own
    # docstring) has somewhere consistent to read from.
    case.steps = [models.TestStep(step_no=s["step_no"], step_text=s.get("step_text"),
                                   expected_result=s.get("expected_result")) for s in steps]
    case.current_draft_version_id = version.id
    _sync_case_mirror(case, version)
    return case


# ---- Folders ----
@router.get("/projects/{project_id}/folders", response_model=List[schemas.TestFolderOut])
def list_folders(project_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _get_project_or_404(db, project_id)
    return db.query(models.TestFolder).filter_by(project_id=project_id).order_by(models.TestFolder.name).all()


@router.post("/projects/{project_id}/folders", response_model=schemas.TestFolderOut)
def create_folder(project_id: int, payload: schemas.TestFolderCreate, db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    name = payload.name.strip()
    if not name:
        raise HTTPException(400, "Folder name cannot be blank")
    if payload.parent_id:
        parent = db.query(models.TestFolder).filter_by(id=payload.parent_id, project_id=project_id).first()
        if not parent:
            raise HTTPException(404, "Parent folder not found in this project")
    obj = models.TestFolder(project_id=project_id, parent_id=payload.parent_id, name=name,
                             created_by_id=current_user.id)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def _descendant_folder_ids(db: Session, folder_id: int) -> set:
    """Every folder id nested anywhere beneath folder_id, at any depth --
    used by _validate_new_parent's cycle guard below (a folder can never be
    re-parented under one of its own descendants, since that would
    disconnect it from the project's tree entirely -- parent_id would point
    at a node that itself hangs off the folder being moved)."""
    ids: set = set()
    frontier = [folder_id]
    while frontier:
        rows = db.query(models.TestFolder.id).filter(models.TestFolder.parent_id.in_(frontier)).all()
        frontier = [row[0] for row in rows]
        ids.update(frontier)
    return ids


def _validate_new_parent(db: Session, folder_id: int, project_id: int, new_parent_id: Optional[int]) -> None:
    """Shared cycle/existence guard for both move_folder and update_folder's
    parent_id reassignment -- raises if new_parent_id isn't a real folder in
    the same project, is the folder itself, or is one of its own descendants
    (either of which would disconnect the folder from the project's tree
    entirely). A None new_parent_id (top level) is always valid."""
    if new_parent_id is None:
        return
    if new_parent_id == folder_id:
        raise HTTPException(400, "A folder cannot be its own parent")
    parent = db.query(models.TestFolder).filter_by(id=new_parent_id, project_id=project_id).first()
    if not parent:
        raise HTTPException(404, "Destination folder not found in this project")
    if new_parent_id in _descendant_folder_ids(db, folder_id):
        raise HTTPException(400, "A folder cannot be moved into one of its own sub-folders")


@router.patch("/folders/{folder_id}", response_model=schemas.TestFolderOut)
def update_folder(folder_id: int, payload: schemas.TestFolderUpdate, db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Reported directly: "Once folder is created, folder details should be
    editable." A folder's only real "detail" beyond its position in the tree
    is its name -- parent_id is accepted here too (same semantics as
    move_folder, via the shared _validate_new_parent guard) purely so this
    general-purpose PATCH is a complete edit endpoint on its own, even though
    the UI's dedicated Move action (see move_folder) is the primary, faster
    path for repositioning a folder."""
    obj = get_or_404(db, models.TestFolder, folder_id, "Folder")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    data = payload.model_dump(exclude_unset=True)
    if "name" in data:
        name = (data["name"] or "").strip()
        if not name:
            raise HTTPException(400, "Folder name cannot be blank")
        obj.name = name
    if "parent_id" in payload.model_fields_set:
        new_parent_id = data.get("parent_id")
        _validate_new_parent(db, folder_id, obj.project_id, new_parent_id)
        obj.parent_id = new_parent_id
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/folders/{folder_id}")
def delete_folder(folder_id: int, db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Folder deletion is a governance action (GOV-002 maker-checker tier).
    2026-08 whole-module simplification: no more project-membership carve-out
    -- this is QA Lead Group-only now (QA_LEAD/CHIEF_MANAGER_QA/AGM_QA), same
    as every other repository-governance action not tied to a specific
    TestCaseVersion's old/new workflow status (see
    require_can_manage_repository_governance in deps.py)."""
    obj = get_or_404(db, models.TestFolder, folder_id, "Folder")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_manage_repository_governance(current_user)
    has_children = db.query(models.TestFolder).filter_by(parent_id=folder_id).first()
    has_cases = db.query(models.TestCase).filter_by(folder_id=folder_id).first()
    if has_children or has_cases:
        raise HTTPException(400, "Move or remove everything inside this folder before deleting it")
    db.delete(obj)
    db.commit()
    return {"ok": True}


@router.post("/folders/{folder_id}/move", response_model=schemas.TestFolderOut)
def move_folder(folder_id: int, payload: schemas.TestFolderMove, db: Session = Depends(get_db),
                 current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Re-parents a folder (and everything already nested beneath it, which
    moves along with it automatically since children only reference their
    parent's id) to a different point in the same project's tree, or to top
    level if parent_id is null."""
    obj = get_or_404(db, models.TestFolder, folder_id, "Folder")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    _validate_new_parent(db, folder_id, obj.project_id, payload.parent_id)
    obj.parent_id = payload.parent_id
    db.commit()
    db.refresh(obj)
    return obj


def _clone_folder_subtree(db: Session, source: models.TestFolder, new_parent_id: Optional[int],
                           current_user: models.User, name_override: Optional[str] = None) -> models.TestFolder:
    """Recursively duplicates `source` -- itself, every test case directly
    inside it (own steps included), and every child folder at any depth,
    each with its own test cases -- under new_parent_id. Every cloned test
    case is a brand-new identity (its own governed TQA-TC key, own first
    Draft version at 1.0) with status Draft, exactly like create_test_case/
    import already do for anything new: a copy is not the same approved
    artifact as its source, so it re-enters QA Lead review rather than
    inheriting the source case's current approval status."""
    new_folder = models.TestFolder(
        project_id=source.project_id, parent_id=new_parent_id,
        name=name_override or source.name, created_by_id=current_user.id,
    )
    db.add(new_folder)
    db.flush()

    for case in db.query(models.TestCase).filter_by(folder_id=source.id).order_by(models.TestCase.test_case_key).all():
        source_version = _current_display_version(case)
        content = {f: getattr(source_version, f) if source_version else getattr(case, f) for f in _CONTENT_FIELDS}
        steps = [{"step_no": s.step_no, "step_text": s.step_text, "expected_result": s.expected_result}
                 for s in (source_version.steps if source_version else case.steps)]
        new_case = _create_case_with_first_draft(
            db, source.project_id, new_folder.id, content, list(case.tags), steps, current_user,
            source_version_id=source_version.id if source_version else None,
        )
        # Reported directly ("100 testcases... not possible to manually
        # submit one by one") -- this used to log "submitted to the QA Lead"
        # while actually leaving the new copy in Draft (submit_test_case was
        # never called), so QA Lead had nothing to approve despite what the
        # activity feed claimed. Now it actually submits whenever the copied
        # steps are complete enough to pass TC-003 (true for anything copied
        # from an already-Approved source); falls back to an honest "still
        # Draft" audit note only for the rare copy with no usable steps.
        new_draft = new_case.current_draft_version
        try:
            _validate_steps(new_draft.steps)
        except HTTPException:
            db.add(_case_workflow_action(
                new_case.id, current_user, "Cloned",
                f"Duplicated from {case.test_case_key} via folder copy; left in Draft because it has no "
                "complete steps to submit for review yet.",
            ))
        else:
            _submit_draft(
                db, new_case, new_draft, current_user, None,
                extra_comment=f"Duplicated from {case.test_case_key} via folder copy and submitted for Reviewer recommendation.",
            )

    children = db.query(models.TestFolder).filter_by(parent_id=source.id).order_by(models.TestFolder.name).all()
    for child in children:
        _clone_folder_subtree(db, child, new_folder.id, current_user)

    return new_folder


@router.post("/folders/{folder_id}/copy", response_model=schemas.TestFolderOut)
def copy_folder(folder_id: int, payload: schemas.TestFolderCopy, db: Session = Depends(get_db),
                 current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Duplicates a folder and its entire subtree -- child folders and every
    test case inside, at any depth -- into a new, independent copy."""
    source = get_or_404(db, models.TestFolder, folder_id, "Folder")
    _require_active_project(_get_project_or_404(db, source.project_id))
    require_can_author_repository(db, source.project_id, current_user)
    dest_parent_id = payload.parent_id if payload.parent_id is not None else source.parent_id
    if dest_parent_id:
        dest_parent = db.query(models.TestFolder).filter_by(id=dest_parent_id, project_id=source.project_id).first()
        if not dest_parent:
            raise HTTPException(404, "Destination folder not found in this project")
    name_override = None
    if payload.name is not None:
        name_override = payload.name.strip()
        if not name_override:
            raise HTTPException(400, "Folder name cannot be blank")
    elif dest_parent_id == source.parent_id:
        name_override = f"{source.name} (Copy)"
    new_root = _clone_folder_subtree(db, source, dest_parent_id, current_user, name_override)
    db.commit()
    db.refresh(new_root)
    return new_root


# ---- Test Cases ----
@router.get("/projects/{project_id}/test-cases", response_model=pagination.Page[schemas.TestCaseListOut])
def list_test_cases(
    project_id: int,
    folder_id: Optional[str] = Query(
        None,
        description="Filter to one folder's direct contents by numeric id, "
                    "the literal 'unfiled' for folder_id IS NULL, or omitted for the whole project",
    ),
    priority: Optional[str] = Query(None, description="Exact-match priority filter"),
    tag: Optional[str] = Query(None, description="Exact-match tag filter"),
    cursor_mode: bool = Query(False, description="Use primary-key cursor pagination instead of OFFSET"),
    cursor: Optional[int] = Query(None, ge=0),
    params: pagination.PageParams = Depends(),
    db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user),
):
    """SRS 7.2 pagination rollout -- was previously "every test case in the
    project", with the Repository UI grouping the complete result into its
    folder tree and doing all filtering/pagination client-side. The folder
    tree itself (GET .../folders below) stays a full, unpaginated fetch --
    it's lightweight and structural, not a growth-unbounded list -- but this
    endpoint (the potentially large one) is now paginated, with folder
    selection, priority, and tag becoming server-side filters instead of an
    in-browser .filter() over the whole project. See TestCaseSummaryOut
    below for the folder-tree counts/tag list/project-wide stats this list
    endpoint no longer has enough data in hand to compute on its own."""
    _get_project_or_404(db, project_id)
    # 2026-08 "Recycle Bin" requirement -- a soft-deleted case is never part
    # of this normal list, regardless of any status/folder/search filter --
    # it only ever shows up via the dedicated Recycle Bin view/endpoint
    # (list_recycle_bin below) until restored or purged.
    q = (
        db.query(models.TestCase)
        .filter(models.TestCase.project_id == project_id, models.TestCase.is_deleted == False)  # noqa: E712 - Oracle requires = 0, not IS 0
        .options(*_LIST_CASE_EAGER_LOADS)
    )
    if folder_id == "unfiled":
        q = q.filter(models.TestCase.folder_id.is_(None))
    elif folder_id:
        try:
            folder_id_int = int(folder_id)
        except ValueError:
            raise HTTPException(400, "folder_id must be numeric or 'unfiled'")
        q = q.filter(models.TestCase.folder_id == folder_id_int)
    if priority:
        q = q.filter(models.TestCase.priority == priority)
    if tag:
        q = q.join(models.TestCaseTag, models.TestCaseTag.test_case_id == models.TestCase.id).filter(models.TestCaseTag.tag == tag)
    q = pagination.apply_search(
        q, params,
        models.TestCase.test_case_key, models.TestCase.test_scenario, models.TestCase.epic_id,
        models.TestCase.cr_number, models.TestCase.feature_id, models.TestCase.user_story_id,
        models.TestCase.module_name,
    )
    q = pagination.apply_status_filter(q, params, models.TestCase.status)
    # 2026-08 -- "Final-Approved Test Case Deletion and Archive Requirement":
    # "Archived test cases are excluded from active lists by default" /
    # "removed from the default active test-case list." Only applies when
    # the caller hasn't asked for any specific status -- an explicit filter
    # (including explicitly filtering FOR "Archived" itself, which is how
    # the existing status dropdown's "Archived" option already satisfies
    # "available under an Archived Test Cases filter") is left completely
    # alone. Deliberately done here rather than by hard-coding the status
    # enumeration on the frontend (frontend/src/constants.ts's
    # TEST_CASE_STATUSES), since that list would need to be kept in perfect
    # sync with every current and future status value or risk silently
    # hiding a real (non-Archived) status from the default view instead.
    if not params.status:
        q = q.filter(models.TestCase.status != "Archived")
    if cursor_mode:
        return pagination.paginate_by_id(q, params, models.TestCase.id, cursor)
    q = pagination.apply_sort(q, params, sortable={
        "test_case_key": models.TestCase.test_case_key,
        "status": models.TestCase.status,
        "priority": models.TestCase.priority,
        "updated_at": models.TestCase.updated_at,
    }, default_column=models.TestCase.created_at, id_column=models.TestCase.id)
    result = pagination.paginate(q, params)
    return pagination.to_page_response(result, params)


@router.get("/projects/{project_id}/test-cases/summary", response_model=schemas.TestCaseSummaryOut)
def get_test_case_summary(project_id: int, db: Session = Depends(get_db),
                           current_user: models.User = Depends(get_current_user)):
    """The folder tree's per-folder/unfiled counts, the tag filter dropdown's
    options, and the "Test cases / Approved / Pending review / Critical"
    stat bar all used to be computed in the browser from the complete
    (unpaginated) project case list -- now that the main list above is
    paginated, none of those has enough data on hand any more. This is the
    one place that recomputes all of them, via SQL COUNT/GROUP BY against
    the whole project regardless of which page/folder/filter the main list
    currently has selected -- never a full-row fetch."""
    _get_project_or_404(db, project_id)
    # 2026-08 "Recycle Bin" requirement -- every count here mirrors what the
    # default (non-Archived, non-deleted) list actually shows; a soft-
    # deleted case is excluded the same way list_test_cases excludes it, and
    # gets its own dedicated recycle_bin_count instead. archived_count is new
    # too, powering the sidebar's "Archived" shortcut badge the same way
    # unfiled_count already powers "Unfiled"'s.
    base = db.query(models.TestCase).filter(
        models.TestCase.project_id == project_id, models.TestCase.is_deleted == False,  # noqa: E712 - Oracle requires = 0, not IS 0
    )
    total = base.filter(models.TestCase.status != "Archived").count()
    archived_count = base.filter(models.TestCase.status == "Archived").count()
    recycle_bin_count = db.query(models.TestCase).filter(
        models.TestCase.project_id == project_id, models.TestCase.is_deleted == True,  # noqa: E712 - Oracle requires = 1, not IS 1
    ).count()
    unfiled_count = base.filter(models.TestCase.folder_id.is_(None), models.TestCase.status != "Archived").count()
    folder_counts = dict(
        db.query(models.TestCase.folder_id, func.count(models.TestCase.id))
        .filter(
            models.TestCase.project_id == project_id, models.TestCase.folder_id.isnot(None),
            models.TestCase.is_deleted == False, models.TestCase.status != "Archived",  # noqa: E712 - Oracle requires = 0, not IS 0
        )
        .group_by(models.TestCase.folder_id).all()
    )
    approved_count = base.filter(models.TestCase.current_approved_version_id.isnot(None)).count()
    # Broadened to also count the NEW-path equivalent status (Recommendation
    # Pending / QA Lead Approval Pending) under the same "Stage 1"/"Stage 2"
    # aggregate -- both vocabularies mean the same thing to this summary
    # card, and TestCaseSummaryOut's field names predate the 2026-08
    # Simplified Test Management change (no schema/frontend change needed).
    in_review_count = base.filter(models.TestCase.status.in_(("In Review", "Recommendation Pending"))).count()
    review_completed_count = base.filter(
        models.TestCase.status.in_(("Review Completed", "QA Lead Approval Pending"))
    ).count()
    critical_count = base.filter(models.TestCase.priority == "Critical").count()
    tags = [
        row[0] for row in db.query(models.TestCaseTag.tag)
        .join(models.TestCase, models.TestCase.id == models.TestCaseTag.test_case_id)
        .filter(models.TestCase.project_id == project_id)
        .distinct().order_by(models.TestCaseTag.tag).all()
    ]
    return schemas.TestCaseSummaryOut(
        total=total, unfiled_count=unfiled_count, folder_counts=folder_counts,
        approved_count=approved_count, in_review_count=in_review_count,
        review_completed_count=review_completed_count, critical_count=critical_count,
        tags=tags, archived_count=archived_count, recycle_bin_count=recycle_bin_count,
    )


@router.get("/projects/{project_id}/test-cases/all", response_model=List[schemas.TestCaseListOut], deprecated=True)
def list_all_test_cases_for_project(project_id: int, db: Session = Depends(get_db),
                                     current_user: models.User = Depends(get_current_user)):
    """PAG-010 -- deliberately NOT paginated, unlike list_test_cases above.
    This exists solely to source bulk-selection candidate pools that
    genuinely need the complete set, not one page of it -- currently
    TestExecution.tsx's "Add Test Cases to Cycle" modal, which needs every
    Approved/non-Archived case in the project (minus whichever are already
    in the target cycle) so "Select all" can mean all of them, and so the
    "N testcases are unavailable" banner counts the real total pending
    review, not just whatever page happened to be loaded. Same eager-loads
    and response shape as the paginated endpoint -- just no page/limit."""
    _get_project_or_404(db, project_id)
    return (
        db.query(models.TestCase)
        .filter(models.TestCase.project_id == project_id, models.TestCase.is_deleted == False)  # noqa: E712 - Oracle requires = 0, not IS 0
        .options(*_LIST_CASE_EAGER_LOADS)
        .order_by(models.TestCase.created_at.desc()).all()
    )


@router.get("/projects/{project_id}/export-xlsx")
def export_test_repository(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Export the complete project repository and its QA review history.

    UI filters deliberately do not limit the export: this is a governed
    snapshot of the selected Project, including every status and every
    retained step and review event.
    """
    project = _get_project_or_404(db, project_id)
    folders = (db.query(models.TestFolder).filter_by(project_id=project_id)
               .order_by(models.TestFolder.name).all())
    cases = (db.query(models.TestCase).filter_by(project_id=project_id)
             .order_by(models.TestCase.test_case_key).all())
    folder_by_id = {folder.id: folder for folder in folders}

    def folder_path(folder_id: Optional[int]) -> str:
        if not folder_id:
            return "Unfiled"
        parts, seen = [], set()
        current = folder_by_id.get(folder_id)
        while current and current.id not in seen:
            seen.add(current.id)
            parts.append(current.name)
            current = folder_by_id.get(current.parent_id)
        return " / ".join(reversed(parts)) or "Unfiled"

    case_ids = [case.id for case in cases]
    actions = []
    if case_ids:
        actions = (db.query(models.ApprovalAction).filter(
            models.ApprovalAction.entity_type == "TEST_CASE",
            models.ApprovalAction.entity_id.in_(case_ids),
        ).order_by(models.ApprovalAction.created_at).all())
    case_key_by_id = {case.id: case.test_case_key for case in cases}

    workbook = new_workbook()
    add_summary_sheet(
        workbook,
        f"{project.project_key} · Test Repository",
        "Complete repository definition and QA Lead review snapshot.",
        [
            ("Project", project.name), ("Project ID", project.project_key),
            ("Department", project.department or "—"),
            ("Project status", "Active" if project.is_active else "Inactive"),
            ("Generated by", current_user.full_name), ("Generated at", models.now()),
        ],
        [
            ("Test cases", len(cases)),
            ("Approved", sum(case.status == "Approved" for case in cases)),
            ("In Review", sum(case.status == "In Review" for case in cases)),
            ("Review Completed", sum(case.status == "Review Completed" for case in cases)),
            ("Recommendation Pending", sum(case.status == "Recommendation Pending" for case in cases)),
            ("QA Lead Approval Pending", sum(case.status == "QA Lead Approval Pending" for case in cases)),
            ("Returned", sum(case.status == "Returned" for case in cases)),
            ("Returned by QA", sum(case.status == "Returned by QA" for case in cases)),
            ("Returned by QA Lead", sum(case.status == "Returned by QA Lead" for case in cases)),
            ("Rejected", sum(case.status == "Rejected" for case in cases)),
            ("Draft", sum(case.status == "Draft" for case in cases)),
            ("Archived", sum(case.status == "Archived" for case in cases)),
            ("Folders", len(folders)),
        ],
    )

    case_headers = [
        "Test Case ID", "Version", "Folder Path", "Epic ID", "CR Number",
        "Feature ID", "User Story ID", "Test Type", "Module", "Scenario",
        "Pre-Condition", "Description", "Priority", "Tags", "Workflow Status",
        "Created By", "Created At", "Updated At", "Step Count",
    ]
    add_table_sheet(
        workbook, "Test Cases", "Test Case Definitions", case_headers,
        [[
            case.test_case_key, case.version, folder_path(case.folder_id), case.epic_id,
            case.cr_number, case.feature_id, case.user_story_id, case.test_type,
            case.module_name, case.test_scenario, case.pre_condition, case.description,
            case.priority, ", ".join(case.tags), case.status, case.created_by_name, case.created_at,
            case.updated_at, len(case.steps),
        ] for case in cases],
        subtitle="One row per reusable testcase. Workflow Status shows its current version's review state.",
        wrap_headers={"Scenario", "Pre-Condition", "Description"},
        date_headers={"Created At", "Updated At"},
        status_headers={"Workflow Status"},
        widths={"Scenario": 38, "Pre-Condition": 38, "Description": 42, "Folder Path": 26},
    )
    add_table_sheet(
        workbook, "Test Steps", "Test Steps", [
            "Test Case ID", "Version", "Step No", "Step", "Expected Result",
        ], [[
            case.test_case_key, case.version, step.step_no, step.step_text,
            step.expected_result,
        ] for case in cases for step in case.steps],
        subtitle="Every retained step of the current version, in execution order.",
        wrap_headers={"Step", "Expected Result"},
        widths={"Step": 52, "Expected Result": 52},
    )
    add_table_sheet(
        workbook, "Review History", "QA Lead Review History", [
            "Test Case ID", "Review Step", "Decision", "Approver", "Role Snapshot",
            "Comments", "Timestamp",
        ], [[
            case_key_by_id.get(action.entity_id, f"Testcase #{action.entity_id}"),
            action.step_name, action.decision, action.actor_name or "System",
            action.actor_role, action.comments, action.created_at,
        ] for action in actions],
        subtitle="Append-only submission, approval, return, and re-approval events.",
        wrap_headers={"Comments"}, date_headers={"Timestamp"}, status_headers={"Decision"},
        widths={"Comments": 54, "Role Snapshot": 28},
    )
    return workbook_response(workbook, f"{project.project_key}_test_repository.xlsx")


@router.post("/projects/{project_id}/export-xlsx/jobs")
def queue_test_repository_export(
    project_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = _get_project_or_404(db, project_id)
    filename = f"{project.project_key}_test_repository.xlsx"
    user_id = current_user.id

    def build(job_id: str):
        with SessionLocal() as worker_db:
            worker_user = worker_db.query(models.User).get(user_id)
            if not worker_user:
                raise RuntimeError("The user who started this export no longer exists")
            jobs.update(job_id, progress=15)
            response = export_test_repository(project_id, worker_db, worker_user)
            return asyncio.run(jobs.save_streaming_response(job_id, response, filename))

    return jobs.enqueue(background_tasks, "TEST_REPOSITORY_EXPORT", user_id, build)


@router.post("/projects/{project_id}/test-cases", response_model=schemas.TestCaseOut)
def create_test_case(project_id: int, payload: schemas.TestCaseCreate, db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    content = {f: getattr(payload, f) for f in _CONTENT_FIELDS}
    steps = [s.model_dump() for s in payload.steps]
    obj = _create_case_with_first_draft(db, project_id, payload.folder_id, content,
                                         _normalize_tags(payload.tags), steps, current_user)
    # 2026-08 -- reported directly: Save on the testcase modal no longer
    # closes it (frontend/TestRepository.tsx), so a freshly created case's
    # checkout state is now actually visible to the person who just made it,
    # not hidden behind an immediate modal close. _create_case_with_first_
    # draft itself doesn't set checked_out_by_id (it's shared with import/
    # clone, which shouldn't auto-lock every case to the importer), so
    # without this, the case they just created would read back as
    # unreserved and force the form straight into "Read-only until
    # reserved" the instant it was saved. Auto-checking it out to its own
    # author here is scoped to this single-case create endpoint only.
    obj.checked_out_by_id = current_user.id
    db.commit()
    db.refresh(obj)
    return obj


@router.get("/test-cases/by-key/{test_case_key}", response_model=schemas.TestCaseOut)
def get_test_case_by_key(test_case_key: str, db: Session = Depends(get_db),
                         current_user: models.User = Depends(get_current_user)):
    """Resolve a globally unique test-case business ID for Global Search.

    This route must remain above `/test-cases/{case_id}` so FastAPI does not
    try to parse the literal `by-key` path segment as an integer case ID.
    """
    normalized_key = test_case_key.strip().upper()
    obj = db.query(models.TestCase).filter_by(test_case_key=normalized_key).first()
    if not obj:
        raise HTTPException(404, f"Test Case {normalized_key} was not found")
    return obj


@router.get("/test-cases/{case_id}", response_model=schemas.TestCaseOut)
def get_test_case(case_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    return obj


# ---- Version history / compare (VER-005) ----
@router.get("/test-cases/{case_id}/versions", response_model=List[schemas.TestCaseVersionSummary])
def list_test_case_versions(case_id: int, db: Session = Depends(get_db),
                            current_user: models.User = Depends(get_current_user)):
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    return (db.query(models.TestCaseVersion).filter_by(test_case_id=case_id)
            .order_by(models.TestCaseVersion.id.desc()).all())


@router.get("/test-cases/{case_id}/versions/{version_id}", response_model=schemas.TestCaseVersionOut)
def get_test_case_version(case_id: int, version_id: int, db: Session = Depends(get_db),
                          current_user: models.User = Depends(get_current_user)):
    version = db.query(models.TestCaseVersion).filter_by(id=version_id, test_case_id=case_id).first()
    if not version:
        raise HTTPException(404, "Test case version not found")
    return version


@router.get("/test-cases/{case_id}/versions-compare", response_model=schemas.TestCaseVersionCompareOut)
def compare_test_case_versions(case_id: int, left: int, right: int, db: Session = Depends(get_db),
                               current_user: models.User = Depends(get_current_user)):
    """SRS VER-005 -- field-level and step-level differences between any two
    versions of the same testcase."""
    left_v = db.query(models.TestCaseVersion).filter_by(id=left, test_case_id=case_id).first()
    right_v = db.query(models.TestCaseVersion).filter_by(id=right, test_case_id=case_id).first()
    if not left_v or not right_v:
        raise HTTPException(404, "One or both versions were not found on this test case")
    field_diffs = {}
    for field in _CONTENT_FIELDS:
        lv, rv = getattr(left_v, field), getattr(right_v, field)
        if lv != rv:
            field_diffs[field] = {"left": lv, "right": rv}
    left_steps = {s.step_no: s for s in left_v.steps}
    right_steps = {s.step_no: s for s in right_v.steps}
    step_diffs = {}
    for step_no in sorted(set(left_steps) | set(right_steps)):
        l, r = left_steps.get(step_no), right_steps.get(step_no)
        l_data = {"step_text": l.step_text, "expected_result": l.expected_result} if l else None
        r_data = {"step_text": r.step_text, "expected_result": r.expected_result} if r else None
        if l_data != r_data:
            step_diffs[step_no] = {"left": l_data, "right": r_data}
    return schemas.TestCaseVersionCompareOut(left=left_v, right=right_v, field_diffs=field_diffs, step_diffs=step_diffs)


def _enforce_checkout_lock(case: models.TestCase, current_user: models.User) -> None:
    """A checked-out test case is locked for editing/deleting to everyone
    except whoever holds the checkout -- Admin always bypasses (same escape
    hatch every other role gate in this router already gives Admin), so a
    lock nobody remembers to release can still be broken by an Administrator
    without anyone touching the database directly."""
    if case.checked_out_by_id and case.checked_out_by_id != current_user.id and not current_user.has_role(Role.ADMIN):
        raise HTTPException(
            423,
            f"{case.test_case_key} is checked out by {case.checked_out_by_name} and locked for editing. "
            "Check it in first, or ask them to.",
        )


@router.post("/test-cases/{case_id}/checkout", response_model=schemas.TestCaseOut)
def checkout_test_case(case_id: int, db: Session = Depends(get_db),
                        current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Reported directly: "check in checkout option should be available for
    testcases, otherwise multiple people can edit at once, if checkout, the
    testcase is locked for editing by that user." Explicit, SharePoint-style
    checkout -- locks a test case to whoever checks it out so nobody else's
    concurrent edit can silently race with theirs. Idempotent for the
    current holder (re-checking out just refreshes the timestamp); rejected
    for anyone else while it's already held."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    if obj.is_deleted:
        raise HTTPException(400, "This test case is in the Recycle Bin -- restore it before editing")
    draft = obj.current_draft_version
    if draft and draft.status in ("In Review", "Review Completed", "Recommendation Pending", "QA Lead Approval Pending"):
        pending_stage = {
            "In Review": "Reviewer recommendation",
            "Review Completed": "QA Lead final approval",
            "Recommendation Pending": "QA recommendation",
            "QA Lead Approval Pending": "QA Lead final approval",
        }[draft.status]
        raise HTTPException(
            409,
            f"This test case is pending {pending_stage}; editing access is locked until a decision is recorded",
        )
    if obj.checked_out_by_id and obj.checked_out_by_id != current_user.id:
        raise HTTPException(423, f"Already checked out by {obj.checked_out_by_name}")
    obj.checked_out_by_id = current_user.id
    obj.checked_out_at = models.now()
    db.commit()
    db.refresh(obj)
    return obj


@router.post("/test-cases/{case_id}/checkout-override", response_model=schemas.TestCaseOut)
def checkout_override(case_id: int, payload: schemas.TestCaseCheckoutOverride, db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """TC-004 "QA Lead and Administrator override shall require a reason and
    audit event" -- forcibly takes the checkout away from whoever currently
    holds it. Unlike checkout_test_case's own soft 423 rejection, this always
    succeeds for the QA Lead Group (QA_LEAD/CHIEF_MANAGER_QA/AGM_QA -- 2026-08
    whole-module simplification dropped the project-membership carve-out, see
    require_can_manage_repository_governance in deps.py), but only when a
    reason is given."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_manage_repository_governance(current_user)
    draft = obj.current_draft_version
    if draft and draft.status in ("In Review", "Review Completed", "Recommendation Pending", "QA Lead Approval Pending"):
        raise HTTPException(409, "Checkout cannot be overridden while this test case is pending an approval decision")
    if not (payload.reason or "").strip():
        raise HTTPException(400, "A reason is required to override an existing checkout")
    previous_holder = obj.checked_out_by_name
    obj.checked_out_by_id = current_user.id
    obj.checked_out_at = models.now()
    db.add(_case_workflow_action(
        obj.id, current_user, "Checkout override",
        f"Checkout forcibly reassigned from {previous_holder or 'nobody'}: {payload.reason.strip()}",
    ))
    db.commit()
    db.refresh(obj)
    return obj


@router.post("/test-cases/{case_id}/checkin", response_model=schemas.TestCaseOut)
def checkin_test_case(case_id: int, db: Session = Depends(get_db),
                       current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Releases a checkout. Only the holder can normally check a case back
    in; Admin may force-release an abandoned lock, same bypass as the edit/
    delete guard itself. Checking in a case that isn't checked out at all is
    a harmless no-op rather than an error."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    if obj.checked_out_by_id and obj.checked_out_by_id != current_user.id and not current_user.has_role(Role.ADMIN):
        raise HTTPException(
            423,
            f"This test case is checked out by {obj.checked_out_by_name} -- only they "
            "(or an Administrator) can check it back in",
        )
    obj.checked_out_by_id = None
    obj.checked_out_at = None
    db.commit()
    db.refresh(obj)
    return obj


@router.patch("/test-cases/{case_id}", response_model=schemas.TestCaseOut)
def update_test_case(case_id: int, payload: schemas.TestCaseUpdate, db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """SRS VER-003 "Editing an approved testcase shall create a new Draft
    version without modifying the approved baseline." If a draft revision is
    already in progress (Draft or Returned -- not yet submitted), its
    content is updated in place; otherwise a brand-new Draft version is
    spun off first. folder_id/tags stay identity-level (TestCase's own
    columns, see its docstring) and are applied regardless of version state.

    2026-08 Approval Workflow refactor -- the new draft is spun off from
    whichever of these is the actual "last known good" content to build on:
    the current draft if it's Rejected (terminal -- frozen in history, not
    edited in place, same mechanic as spinning off an Approved baseline),
    else the current approved baseline, else nothing (a case that somehow
    has neither -- shouldn't happen post-creation, but never crash an edit
    over it). In Review/Review Completed stay blocked from any edit --
    there's a decision pending, whichever stage it's at."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    _enforce_checkout_lock(obj, current_user)
    if obj.is_deleted:
        raise HTTPException(400, "This test case is in the Recycle Bin -- restore it before editing")
    if "status" in payload.model_fields_set:
        raise HTTPException(400, "Test case status is controlled by the review workflow and cannot be set directly")

    draft = obj.current_draft_version
    if draft and draft.status == "In Review":
        raise HTTPException(400, "This test case is pending Reviewer recommendation -- wait for a decision before editing again")
    if draft and draft.status == "Review Completed":
        raise HTTPException(400, "This test case is pending QA Lead final approval -- wait for a decision before editing again")
    if draft and draft.status == "Recommendation Pending":
        raise HTTPException(400, "This test case is pending QA recommendation -- wait for a decision before editing again")
    if draft and draft.status == "QA Lead Approval Pending":
        raise HTTPException(400, "This test case is pending QA Lead final approval -- wait for a decision before editing again")
    # 2026-08 -- "Final-Approved Test Case Deletion and Archive Requirement":
    # "Not be editable while archived." Without a draft in progress, editing
    # would otherwise spin off a new draft from `obj.current_approved_version`
    # below regardless of ITS status -- including an Archived one, silently
    # bringing an archived case back into active editing without ever going
    # through Restore. Blocked explicitly here instead.
    if not draft and obj.current_approved_version and obj.current_approved_version.status == "Archived":
        raise HTTPException(400, "This test case is archived -- restore it before editing")
    rejected_base = draft if (draft and draft.status == "Rejected") else None
    if not draft or rejected_base:
        # VER-003 (extended for Rejected, see docstring above): spin off a
        # new draft from the best available content base.
        base = obj.current_approved_version or rejected_base
        base_content = {f: getattr(base, f) if base else None for f in _CONTENT_FIELDS}
        major, minor = _next_provisional_version_numbers(obj)
        draft = models.TestCaseVersion(
            test_case_id=obj.id, version_major=major, version_minor=minor, status="Draft",
            author_id=current_user.id, created_at=models.now(),
            source_version_id=base.id if base else None, **base_content,
        )
        db.add(draft)
        db.flush()
        if base:
            draft.steps = [models.TestCaseVersionStep(step_no=s.step_no, step_text=s.step_text,
                                                       expected_result=s.expected_result) for s in base.steps]
        obj.current_draft_version_id = draft.id
        if rejected_base:
            db.add(_case_workflow_action(
                obj.id, current_user, "Revised after rejection",
                f"Started a new draft (v{draft.version}) off the rejected v{rejected_base.version} for correction.",
                previous_state="Rejected", new_state="Draft",
            ))

    data = payload.model_dump(exclude_unset=True, exclude={"steps", "tags", "folder_id"})
    for field, value in data.items():
        if field in _CONTENT_FIELDS:
            setattr(draft, field, value)
    if payload.steps is not None:
        _replace_draft_steps(db, obj, draft, payload.steps)
    if "folder_id" in payload.model_fields_set:
        obj.folder_id = payload.folder_id
    if payload.tags is not None:
        obj.tag_rows = [models.TestCaseTag(tag=tag) for tag in _normalize_tags(payload.tags)]
    _sync_case_mirror(obj, draft)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/test-cases/{case_id}", response_model=schemas.TestCaseOut)
def delete_test_case(case_id: int, db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """A testcase that has ever been approved OR rejected is governed
    history -- Archive it instead (TC-006 preserves versions/cycle
    membership/execution history while blocking new selection). Delete is
    only allowed while it has never left Draft/Returned, i.e. it was never
    actually decided on by anyone yet. Section 11 "Rejected and superseded
    versions shall remain readable to authorized users for traceability" --
    a Rejected version must never be able to vanish via delete, same as an
    Approved one.

    2026-08 "Recycle Bin" requirement -- "any delete testcases before
    approve will go to recycle bin. only QA lead can clear from recycle
    bin." This no longer issues a real `db.delete()` -- it soft-deletes
    (is_deleted=True) instead, excluding the case from every normal list/
    summary query while keeping the row (and its full version/step history)
    intact and recoverable via restore_test_case_from_recycle_bin. Only
    purge_test_case/bulk_purge_test_cases (QA Lead Group only) still perform
    a real, irreversible `db.delete()`."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    _enforce_checkout_lock(obj, current_user)
    if obj.is_deleted:
        raise HTTPException(400, "This test case is already in the Recycle Bin")
    ever_decided = db.query(models.TestCaseVersion.id).filter(
        models.TestCaseVersion.test_case_id == case_id,
        models.TestCaseVersion.status.in_(("Approved", "Archived", "Rejected")),
    ).first()
    if ever_decided:
        # 409 Conflict, not 400 -- this isn't a malformed request, it's a
        # request that's individually well-formed but conflicts with the
        # test case's own governed state (it's a controlled asset once
        # decided on). `detail` stays a plain string per this app's own
        # Section 8 API standard (main.py::http_exception_handler) --
        # frontend error parsing depends on it -- rather than the nested
        # {"message":..., "test_case_id":...} shape a pasted external spec
        # asked for.
        raise HTTPException(
            409,
            f"Test case {obj.test_case_key} has an approval/rejection history and cannot be permanently "
            f"deleted -- archive it instead.",
        )
    obj.is_deleted = True
    obj.deleted_by_id = current_user.id
    obj.deleted_at = models.now()
    db.add(_case_workflow_action(obj.id, current_user, "Moved to Recycle Bin",
                                 "Deleted before approval -- recoverable from the Recycle Bin until purged."))
    db.commit()
    db.refresh(obj)
    return obj


def _validate_stage2_assignee(db: Session, drafts: List[models.TestCaseVersion],
                              qa_lead_id: int, prohibited_ids: Optional[set] = None) -> models.User:
    qa_lead = db.query(models.User).get(qa_lead_id)
    if not qa_lead or not qa_lead.is_active:
        raise HTTPException(400, "Select an active QA Lead (Stage 2)")
    if not qa_lead.has_department(*TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS):
        raise HTTPException(
            400,
            f"Selected approvers must be mapped to one of: {', '.join(TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS)}",
        )
    # Bug found via debugging pass: this rejected AGM_QA outright, directly
    # contradicting _stage2_approver_ids just above ("All active CM QA and
    # AGM QA users; either one may complete Stage 2") -- an AGM_QA-only user
    # could be picked in the UI (eligible-users isn't role-filtered) and get
    # a 400 here on save. AGM_QA now accepted here too, matching the rest of
    # this file's own Stage 2 semantics.
    if not qa_lead.has_role(Role.QA_LEAD, Role.CHIEF_MANAGER_QA, Role.AGM_QA):
        raise HTTPException(400, "Stage 2 must be assigned to a QA Lead, Chief Manager - QA, or AGM - QA")
    blocked_ids = {draft.author_id for draft in drafts}
    blocked_ids.update(prohibited_ids or set())
    if qa_lead_id in blocked_ids:
        raise HTTPException(400, "The Stage 2 QA Lead must be different from the testcase author and Stage 1 Reviewer")
    return qa_lead


def _validate_submission_assignees(db: Session, drafts: List[models.TestCaseVersion],
                                   reviewer_id: int, qa_lead_id: int) -> None:
    reviewer = db.query(models.User).get(reviewer_id)
    if not reviewer or not reviewer.is_active:
        raise HTTPException(400, "Select an active Reviewer (Stage 1)")
    if not reviewer.has_department(*TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS):
        raise HTTPException(
            400,
            f"Selected approvers must be mapped to one of: {', '.join(TEST_MANAGEMENT_ELIGIBLE_DEPARTMENTS)}",
        )
    if reviewer_id in {draft.author_id for draft in drafts}:
        raise HTTPException(400, "A testcase author cannot also be selected as its Stage 1 Reviewer")
    _validate_stage2_assignee(db, drafts, qa_lead_id, prohibited_ids={reviewer_id})


def _submit_draft(db: Session, case: models.TestCase, draft: models.TestCaseVersion,
                  current_user: models.User, note: Optional[str], extra_comment: Optional[str] = None) -> None:
    """Shared by submit_test_case (single) and bulk_submit_test_cases below --
    transitions one Draft/Returned version onward, releases any checkout,
    syncs the mirror, and writes the audit event. Caller has already
    validated the draft is ready (status check + _validate_steps).

    Routes to whichever workflow this specific draft belongs to (see
    _OLD_WORKFLOW_STATUSES above): a draft already returned under the OLD,
    generic "Returned" status restarts at OLD-path "In Review" exactly as
    before. Everything else -- a fresh Draft never submitted, or a NEW-path
    "Returned by QA"/"Returned by QA Lead" resubmitting -- moves to NEW-path
    "Recommendation Pending" and is routed to the QA Group, not an
    individually-assigned Reviewer (2026-08 "Simplified Test Management
    Review and Approval" requirement)."""
    was_returned = draft.status in ("Returned", "Returned by QA", "Returned by QA Lead")
    previous_state = draft.status
    is_old_path = previous_state == "Returned"
    draft.status = "In Review" if is_old_path else "Recommendation Pending"
    draft.submitted_by_id = current_user.id
    draft.submitted_at = models.now()
    draft.submit_note = (note or "").strip() or None
    # A returned version always restarts at Stage 1. Historical decisions
    # remain in ApprovalAction; these fields describe only the current pass.
    draft.reviewed_by_id = None
    draft.reviewed_at = None
    draft.review_comments = None
    draft.qa_lead_decided_by_id = None
    draft.qa_lead_decided_at = None
    draft.qa_lead_decision_comments = None
    # Group routing is authoritative -- no individually-assigned Reviewer/QA
    # Lead in either workflow's submission path (TM's "4.3 No Reviewer
    # Selection"). Legacy per-version assignments are cleared so they cannot
    # strand a review with an unavailable employee.
    draft.assigned_reviewer_id = None
    draft.assigned_qa_lead_id = None
    case.checked_out_by_id = None
    case.checked_out_at = None
    _sync_case_mirror(case, draft)
    db.add(_case_workflow_action(
        case.id, current_user, "Resubmitted for review" if was_returned else "Submitted for review",
        draft.submit_note or extra_comment or (
            "Resubmitted for QA recommendation." if was_returned
            else "Submitted for QA recommendation."
        ),
        previous_state=previous_state, new_state=draft.status,
    ))


@router.post("/test-cases/{case_id}/submit", response_model=schemas.TestCaseOut)
def submit_test_case(case_id: int, payload: schemas.TestCaseSubmit, db: Session = Depends(get_db),
                     current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """REV-001 -- submits the current Draft/Rework-Required version for QA
    Lead review. Validates completeness (TC-003), releases the checkout,
    and records author/timestamp/optional note."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    draft = obj.current_draft_version
    if not draft or draft.status not in ("Draft", "Returned", "Returned by QA", "Returned by QA Lead"):
        raise HTTPException(400, "There is no draft revision ready to submit for review")
    _validate_steps(draft.steps)
    eligible = (
        _stage1_reviewer_ids(db, obj.project_id, draft.author_id) if draft.status == "Returned"
        else _qa_group_ids(db, draft.author_id)
    )
    if not eligible:
        raise HTTPException(400, "No eligible QA reviewer is configured -- assign at least one active QA_ENGINEER")
    _submit_draft(db, obj, draft, current_user, payload.note)
    db.commit()
    db.refresh(obj)
    return obj


@router.post("/projects/{project_id}/test-cases/bulk-submit", response_model=List[schemas.TestCaseOut])
def bulk_submit_test_cases(project_id: int, payload: schemas.TestCaseBulkSubmit,
                           db: Session = Depends(get_db),
                           current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """REV-001, bulk form. Reported directly: "i have 100 testcases, so it
    is not possible to do manually edit one by one by author to submit" --
    submits every selected case's current Draft/Rework-Required version for
    Reviewer recommendation in one action. Validation (draft exists in the right
    state, TC-003 step completeness) runs for the WHOLE selection before any
    case is changed, so one incomplete case doesn't leave the batch
    half-submitted -- same all-or-nothing convention as
    bulk_approve_test_cases/bulk_delete_test_cases above. Does NOT require a
    checkout (matching submit_test_case's own single-case behavior, which
    has never required one either -- submitting simply clears any existing
    checkout as part of the same action)."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    rows = _selected_project_cases(db, project_id, payload.ids)
    not_ready = [row.test_case_key for row in rows
                 if not row.current_draft_version
                 or row.current_draft_version.status not in
                 ("Draft", "Returned", "Returned by QA", "Returned by QA Lead")]
    if not_ready:
        preview = ", ".join(not_ready[:5])
        suffix = "…" if len(not_ready) > 5 else ""
        raise HTTPException(
            400,
            f"{len(not_ready)} selected test case(s) have no draft revision ready to submit "
            f"(already In Review, Approved, or Archived): {preview}{suffix}",
        )
    step_errors = []
    for row in rows:
        try:
            _validate_steps(row.current_draft_version.steps)
        except HTTPException as exc:
            step_errors.append(f"{row.test_case_key}: {exc.detail}")
    if step_errors:
        preview = "; ".join(step_errors[:5])
        suffix = "…" if len(step_errors) > 5 else ""
        raise HTTPException(400, f"Fix these before submitting: {preview}{suffix}")
    unavailable = [row.test_case_key for row in rows if not (
        _stage1_reviewer_ids(db, project_id, row.current_draft_version.author_id)
        if row.current_draft_version.status == "Returned"
        else _qa_group_ids(db, row.current_draft_version.author_id)
    )]
    if unavailable:
        raise HTTPException(400, "No eligible QA reviewer is configured for: " + ", ".join(unavailable[:5]))
    for row in rows:
        _submit_draft(db, row, row.current_draft_version, current_user, payload.note,
                      extra_comment="Submitted to the shared QA reviewer queue (bulk submit).")
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.post("/test-cases/{case_id}/review", response_model=schemas.TestCaseOut)
def review_test_case(case_id: int, payload: schemas.TestCaseReview, db: Session = Depends(get_db),
                     current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """One endpoint, four possible stages, gated and validated differently
    depending on which stage the target draft is currently sitting in (see
    schemas.TestCaseReview's own docstring for the full decision vocabulary
    per stage). OLD-path drafts already sitting at "In Review"/"Review
    Completed" when the 2026-08 "Simplified Test Management Review and
    Approval" requirement shipped keep running the pre-existing "Test
    Approval Workflow" refactor logic below, completely unchanged:
      "In Review"        -- Reviewer-tier acts (system QA_LEAD role):
                             RECOMMEND -> "Review Completed"; RETURN -> "Returned" (comment mandatory).
      "Review Completed" -- QA-Lead-tier acts (require_can_give_final_approval,
                             CM QA/AGM QA only, deliberately narrower than
                             Reviewer-tier -- see its own docstring in
                             deps.py): APPROVE -> "Approved" (the only
                             decision that activates the version for
                             cycles); RETURN -> "Returned" (comment mandatory);
                             REJECT -> "Rejected", terminal (comment mandatory).
    NEW-path drafts (any fresh Draft submission, or a NEW-vocabulary Returned
    status resubmitting) instead route to the QA Group / QA Lead Group --
    see TEST_CASE_NEW_STATUSES in constants.py and ORACLE_MIGRATION_2026-07.md
    for the reported requirement and the "new cases only" migration decision:
      "Recommendation Pending"   -- any active QA_ENGINEER ("QA Group"),
                                    excluding the author: RECOMMEND ->
                                    "QA Lead Approval Pending"; RETURN ->
                                    "Returned by QA" (comment mandatory);
                                    REJECT -> "Rejected", terminal (comment
                                    mandatory).
      "QA Lead Approval Pending" -- any active member of the QA Lead Group
                                    (QA_LEAD/CHIEF_MANAGER_QA/AGM_QA),
                                    excluding the author: APPROVE ->
                                    "Approved"; RETURN -> "Returned by QA
                                    Lead" (comment mandatory); REJECT ->
                                    "Rejected", terminal (comment mandatory).
    GOV-002 maker-checker: the author of THIS draft version may not act on
    it at any stage, regardless of role -- section 11 "Emergency
    self-approval shall be disabled by default," and this app doesn't build
    the optional override, so it's simply always disabled, no exceptions."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    draft_id = obj.current_draft_version_id
    draft = (
        db.query(models.TestCaseVersion)
        .filter(models.TestCaseVersion.id == draft_id)
        .populate_existing()
        .with_for_update()
        .one_or_none()
    ) if draft_id else None
    _pending_statuses = ("In Review", "Review Completed", "Recommendation Pending", "QA Lead Approval Pending")
    if not draft or draft.status not in _pending_statuses:
        raise HTTPException(409, "This review or approval has already been completed by another authorized user")
    if draft.author_id == current_user.id:
        raise HTTPException(403, "GOV-002: the author of a draft version may not act on their own work")
    # 2026-08 "Simplified Test Management" GOV-002 gap closed -- reported
    # directly: Tester 2 (not the draft's author) submitted Tester 1's
    # draft, then Tester 2 was immediately able to record the Stage 1
    # decision on the very item they'd just submitted -- the check above
    # only ever looked at draft.author_id (who wrote the CONTENT), never at
    # who actually performed the submit action or an earlier stage's
    # decision. Maker-checker means nobody acts twice on the same draft's
    # forward progress, regardless of whether they authored its content.
    # NEW-path only (see _OLD_WORKFLOW_STATUSES above) -- OLD-path GOV-002
    # stays exactly the single author_id check it always was.
    if draft.status == "Recommendation Pending" and draft.submitted_by_id == current_user.id:
        raise HTTPException(403, "GOV-002: you submitted this draft for review and cannot record its Stage 1 decision")
    if draft.status == "QA Lead Approval Pending" and current_user.id in (draft.submitted_by_id, draft.reviewed_by_id):
        raise HTTPException(
            403, "GOV-002: you already acted on this draft at an earlier stage and cannot record its Stage 2 decision"
        )
    decision = payload.decision.strip().upper()
    comments = (payload.comments or "").strip()
    previous_state = draft.status

    if draft.status == "Recommendation Pending":
        if not current_user.has_role(Role.QA_ENGINEER):
            raise HTTPException(403, "Stage 1 recommendation is available only to the QA Group")
        if decision not in {"RECOMMEND", "RETURN", "REJECT"}:
            raise HTTPException(400, "Decision must be RECOMMEND, RETURN, or REJECT while pending QA recommendation")
        if decision == "RECOMMEND":
            management_ids = _qa_lead_group_ids(db, draft.author_id)
            if not management_ids:
                raise HTTPException(400, "No active QA Lead Group approver is configured")
            draft.status = "QA Lead Approval Pending"
            draft.reviewed_by_id = current_user.id
            draft.reviewed_at = models.now()
            draft.review_comments = comments or "Recommended for QA Lead final approval."
            _sync_case_mirror(obj, draft)
            action = "Recommended for approval"
            comments = draft.review_comments
        elif decision == "RETURN":
            if not comments:
                raise HTTPException(400, "A reason is required when returning a test case for changes")
            draft.status = "Returned by QA"
            draft.reviewed_by_id = current_user.id
            draft.reviewed_at = models.now()
            draft.review_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Returned for correction"
        else:  # REJECT
            if not comments:
                raise HTTPException(400, "A reason is required when rejecting a test case")
            draft.status = "Rejected"
            draft.reviewed_by_id = current_user.id
            draft.reviewed_at = models.now()
            draft.review_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Rejected"
    elif draft.status == "QA Lead Approval Pending":
        if not current_user.has_role(Role.QA_LEAD, Role.CHIEF_MANAGER_QA, Role.AGM_QA):
            raise HTTPException(403, "Final approval is available only to the QA Lead Group")
        if decision not in {"APPROVE", "RETURN", "REJECT"}:
            raise HTTPException(400, "Decision must be APPROVE, RETURN, or REJECT while pending QA Lead approval")
        if decision == "APPROVE":
            bump = (payload.version_bump or "minor").strip().lower()
            if bump not in ("minor", "major"):
                raise HTTPException(400, "version_bump must be 'minor' or 'major'")
            if bump == "major":
                if not comments:
                    raise HTTPException(400, "A justification is required to approve a major version increment")
                approved_before = obj.current_approved_version
                draft.version_major = (approved_before.version_major + 1) if approved_before else 1
                draft.version_minor = 0
            draft.status = "Approved"
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments or "Approved and activated by QA Lead."
            obj.current_approved_version_id = draft.id
            obj.current_draft_version_id = None
            _sync_case_mirror(obj, draft)
            action = "Approved & Activated"
            comments = draft.qa_lead_decision_comments
        elif decision == "RETURN":
            if not comments:
                raise HTTPException(400, "A reason is required when returning a test case for changes")
            draft.status = "Returned by QA Lead"
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Returned for correction"
        else:  # REJECT
            if not comments:
                raise HTTPException(400, "A reason is required when rejecting a test case")
            draft.status = "Rejected"
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Rejected"
    elif draft.status == "In Review":
        if not current_user.has_role(Role.QA_LEAD):
            raise HTTPException(403, "Stage 1 review is available only to the QA Lead group")
        if decision not in {"RECOMMEND", "APPROVE", "RETURN", "REJECT"}:
            raise HTTPException(400, "Decision must be APPROVE, RETURN, or REJECT while pending Stage 1 QA review")
        if decision in {"RECOMMEND", "APPROVE"}:
            management_ids = _stage2_approver_ids(db, draft.author_id)
            if not management_ids:
                raise HTTPException(400, "No active CM QA or AGM QA approver is configured")
            draft.status = "Review Completed"
            draft.reviewed_by_id = current_user.id
            draft.reviewed_at = models.now()
            draft.review_comments = comments or "Recommended for QA Lead final approval."
            _sync_case_mirror(obj, draft)
            action = "Recommended for approval"
            comments = draft.review_comments
        elif decision == "RETURN":
            if not comments:
                raise HTTPException(400, "A reason is required when returning a test case for changes")
            draft.status = "Returned"
            draft.reviewed_by_id = current_user.id
            draft.reviewed_at = models.now()
            draft.review_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Returned for correction"
        else:
            if not comments:
                raise HTTPException(400, "A reason is required when rejecting a test case")
            draft.status = "Rejected"
            draft.reviewed_by_id = current_user.id
            draft.reviewed_at = models.now()
            draft.review_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Rejected"
    else:  # "Review Completed" (draft.status not in the three branches above)
        require_can_give_final_approval(db, obj.project_id, current_user)
        if decision not in {"APPROVE", "RETURN", "REJECT"}:
            raise HTTPException(400, "Decision must be APPROVE, RETURN, or REJECT while pending QA management approval")
        if decision == "APPROVE":
            bump = (payload.version_bump or "minor").strip().lower()
            if bump not in ("minor", "major"):
                raise HTTPException(400, "version_bump must be 'minor' or 'major'")
            if bump == "major":
                if not comments:
                    raise HTTPException(400, "A justification is required to approve a major version increment")
                approved_before = obj.current_approved_version
                draft.version_major = (approved_before.version_major + 1) if approved_before else 1
                draft.version_minor = 0
            draft.status = "Approved"
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments or "Approved and activated by QA Lead."
            obj.current_approved_version_id = draft.id
            obj.current_draft_version_id = None
            _sync_case_mirror(obj, draft)
            action = "Approved & Activated"
            comments = draft.qa_lead_decision_comments
        elif decision == "RETURN":
            if not comments:
                raise HTTPException(400, "A reason is required when returning a test case for changes")
            draft.status = "Returned"
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Returned for correction"
        else:  # REJECT
            if not comments:
                raise HTTPException(400, "A reason is required when rejecting a test case")
            draft.status = "Rejected"
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments
            _sync_case_mirror(obj, draft)
            action = "Rejected"
    db.add(_case_workflow_action(obj.id, current_user, action, comments,
                                 previous_state=previous_state, new_state=draft.status))
    db.commit()
    db.refresh(obj)
    return obj


@router.patch("/test-cases/{case_id}/approvers", response_model=schemas.TestCaseOut)
def reassign_test_case_approvers(case_id: int, payload: schemas.TestCaseReassignApprovers,
                                 db: Session = Depends(get_db),
                                 current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Retained temporarily for older clients; individual routing is disabled."""
    raise HTTPException(409, "Individual reviewer assignment is disabled; test cases use automatic role-based group routing")


@router.post("/test-cases/{case_id}/archive", response_model=schemas.TestCaseOut)
def archive_test_case(case_id: int, payload: schemas.TestCaseArchive, db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """TC-006 -- archives the approved baseline: preserves all versions,
    cycle membership, and execution history while preventing new cycle
    selection (add-to-cycle in test_execution.py only accepts Approved
    versions). Any in-progress draft revision is left untouched and stays
    editable -- archiving only retires the approved baseline itself.
    2026-08 whole-module simplification: QA Lead Group-only now, no project-
    membership carve-out (require_can_manage_repository_governance)."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_manage_repository_governance(current_user)
    approved = obj.current_approved_version
    if not approved:
        raise HTTPException(400, "This test case has no approved version to archive")
    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(400, "An archive reason is required")
    approved.status = "Archived"
    if not obj.current_draft_version:
        _sync_case_mirror(obj, approved)
    db.add(_case_workflow_action(
        obj.id, current_user, "Archived", reason,
        previous_state="Approved", new_state="Archived",
    ))
    db.commit()
    db.refresh(obj)
    return obj


@router.post("/test-cases/{case_id}/restore", response_model=schemas.TestCaseOut)
def restore_test_case(case_id: int, db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Appendix A: Archived -> "optionally restore through approval" ->
    Approved. Restores the archived baseline back to Approved.
    2026-08 whole-module simplification: QA Lead Group-only now, no project-
    membership carve-out (require_can_manage_repository_governance)."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_manage_repository_governance(current_user)
    approved = obj.current_approved_version
    if not approved or approved.status != "Archived":
        raise HTTPException(400, "This test case's approved version is not archived")
    approved.status = "Approved"
    if not obj.current_draft_version:
        _sync_case_mirror(obj, approved)
    db.add(_case_workflow_action(
        obj.id, current_user, "Restored", "Archived version restored to Approved.",
        previous_state="Archived", new_state="Approved",
    ))
    db.commit()
    db.refresh(obj)
    return obj


@router.post("/projects/{project_id}/test-cases/bulk-archive", response_model=List[schemas.TestCaseOut])
def bulk_archive_test_cases(project_id: int, payload: schemas.TestCaseBulkArchive,
                            db: Session = Depends(get_db),
                            current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """2026-08 -- "Final-Approved Test Case Deletion and Archive
    Requirement": bulk counterpart to archive_test_case, alongside
    bulk_delete_test_cases -- lets an authorized user archive several
    governed (Approved) test cases in one action instead of one at a time,
    same "N eligible" pattern the frontend already surfaces for bulk-delete.
    Rows already Archived are silently skipped rather than failing the
    batch (re-archiving isn't a meaningful conflict); rows with no approved
    baseline at all are rejected the same way the single-case endpoint
    rejects them, since selecting one for "Archive" in the first place
    would be a frontend eligibility bug."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_manage_repository_governance(current_user)
    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(400, "An archive reason is required")
    rows = _selected_project_cases(db, project_id, payload.ids)
    no_baseline = [row.test_case_key for row in rows
                   if not row.current_approved_version or row.current_approved_version.status not in ("Approved", "Archived")]
    if no_baseline:
        preview = ", ".join(no_baseline[:5])
        suffix = "…" if len(no_baseline) > 5 else ""
        raise HTTPException(
            400,
            f"{len(no_baseline)} selected test case(s) have no approved version to archive: {preview}{suffix}",
        )
    archived_rows = [row for row in rows if row.current_approved_version.status == "Approved"]
    for row in archived_rows:
        approved = row.current_approved_version
        approved.status = "Archived"
        if not row.current_draft_version:
            _sync_case_mirror(row, approved)
        db.add(_case_workflow_action(row.id, current_user, "Archived", reason,
                                     previous_state="Approved", new_state="Archived"))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.post("/projects/{project_id}/test-cases/bulk-restore-from-archive", response_model=List[schemas.TestCaseOut])
def bulk_restore_test_cases(project_id: int, payload: schemas.TestCaseBulkRestoreFromArchive,
                            db: Session = Depends(get_db),
                            current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """2026-08 -- bulk counterpart to restore_test_case, alongside
    bulk_archive_test_cases -- lets an authorized user restore several
    Archived test cases back to Approved in one action instead of one at a
    time, same "N eligible" pattern as bulk-archive/bulk-delete. Rows not
    currently Archived are silently skipped rather than failing the whole
    batch (mirrors bulk-archive's own "already Archived isn't a meaningful
    conflict" reasoning, just in the opposite direction) -- a row selected
    from the Archive view that got restored by someone else a moment
    earlier shouldn't block the rest of the batch."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_manage_repository_governance(current_user)
    rows = _selected_project_cases(db, project_id, payload.ids)
    restorable_rows = [row for row in rows
                       if row.current_approved_version and row.current_approved_version.status == "Archived"]
    for row in restorable_rows:
        approved = row.current_approved_version
        approved.status = "Approved"
        if not row.current_draft_version:
            _sync_case_mirror(row, approved)
        db.add(_case_workflow_action(
            row.id, current_user, "Restored", "Archived version restored to Approved.",
            previous_state="Archived", new_state="Approved",
        ))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.post("/test-cases/{case_id}/clone", response_model=schemas.TestCaseOut)
def clone_test_case(case_id: int, payload: schemas.TestCaseCloneIn, db: Session = Depends(get_db),
                    current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """TC-005 -- creates a NEW testcase identity at version 1.0 Draft,
    recording the source testcase/version it was cloned from. Defaults to
    the same project/folder as the source; either may be overridden
    (project reuse across authorized projects, PRJ-006)."""
    source = get_or_404(db, models.TestCase, case_id, "Test Case")
    target_project_id = payload.project_id or source.project_id
    target_project = _get_project_or_404(db, target_project_id)
    _require_active_project(target_project)
    require_can_author_repository(db, target_project_id, current_user)
    target_folder_id = payload.folder_id if payload.folder_id is not None else (
        source.folder_id if target_project_id == source.project_id else None
    )
    if target_folder_id:
        if not db.query(models.TestFolder).filter_by(id=target_folder_id, project_id=target_project_id).first():
            raise HTTPException(404, "Destination folder not found in that project")
    source_version = _current_display_version(source)
    if not source_version:
        raise HTTPException(400, "This test case has no version to clone")
    content = {f: getattr(source_version, f) for f in _CONTENT_FIELDS}
    if payload.name_suffix:
        content["test_scenario"] = f"{content.get('test_scenario') or ''} {payload.name_suffix}".strip()
    steps = [{"step_no": s.step_no, "step_text": s.step_text, "expected_result": s.expected_result}
             for s in source_version.steps]
    new_case = _create_case_with_first_draft(
        db, target_project_id, target_folder_id, content, list(source.tags), steps, current_user,
        source_version_id=source_version.id,
    )
    db.add(_case_workflow_action(
        new_case.id, current_user, "Cloned",
        f"Cloned from {source.test_case_key} v{source_version.version}.",
    ))
    db.commit()
    db.refresh(new_case)
    return new_case


def _selected_project_cases(db: Session, project_id: int, ids: List[int],
                            include_deleted: bool = False) -> List[models.TestCase]:
    """2026-08 "Recycle Bin" requirement -- `include_deleted` defaults False
    so every existing bulk endpoint (approve/recommend/return/reject/update/
    delete/archive) automatically, uniformly refuses to act on a
    soft-deleted row -- it's treated as not found, same as if it had really
    been deleted, closing off a direct-API path to bypass the Recycle Bin.
    Only bulk_restore_test_cases_from_recycle_bin/bulk_purge_test_cases pass
    include_deleted=True, since operating on a deleted row is the entire
    point of those two."""
    unique_ids = list(dict.fromkeys(ids))
    if not unique_ids:
        raise HTTPException(400, "Select at least one test case")
    q = db.query(models.TestCase).filter(models.TestCase.project_id == project_id, models.TestCase.id.in_(unique_ids))
    if not include_deleted:
        q = q.filter(models.TestCase.is_deleted == False)  # noqa: E712 - Oracle requires = 0, not IS 0
    rows = q.all()
    found = {row.id for row in rows}
    missing = [case_id for case_id in unique_ids if case_id not in found]
    if missing:
        raise HTTPException(404, f"{len(missing)} selected test case(s) were not found in this project")
    return rows


def _normalize_tags(tags: List[str]) -> List[str]:
    normalized = []
    seen = set()
    for raw in tags or []:
        tag = " ".join((raw or "").strip().split())[:80]
        key = tag.casefold()
        if tag and key not in seen:
            seen.add(key)
            normalized.append(tag)
    return normalized[:30]


@router.post("/projects/{project_id}/test-cases/bulk-update", response_model=List[schemas.TestCaseOut])
def bulk_update_test_cases(project_id: int, payload: schemas.TestCaseBulkUpdate,
                           db: Session = Depends(get_db),
                           current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Bulk classification changes (priority/test type/module/tags/folder)
    apply to each selected case's CURRENT draft version if one is in
    progress, else spin off a new draft from the approved baseline for
    each -- same VER-003 rule update_test_case follows for a single case."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    rows = _selected_project_cases(db, project_id, payload.ids)
    changes = payload.model_fields_set - {"ids"}
    if not changes:
        raise HTTPException(400, "Choose at least one field to update")
    if "status" in changes:
        raise HTTPException(400, "Test case status is controlled by the review workflow and cannot be bulk updated")
    testcase_field_changes = changes.intersection({"folder_id", "priority", "test_type", "module_name", "tags"})
    # Routing is deliberately not treated as content editing. It remains
    # available while another user has the case checked out and while the
    # governed review workflow has locked testcase content.
    if testcase_field_changes:
        for row in rows:
            _enforce_checkout_lock(row, current_user)
    workflow_locked_rows = [
        row for row in rows
        if row.status in {"In Review", "Review Completed", "Recommendation Pending", "QA Lead Approval Pending"}
    ]
    if testcase_field_changes and workflow_locked_rows:
        locked_keys = ", ".join(row.test_case_key for row in workflow_locked_rows[:5])
        if len(workflow_locked_rows) > 5:
            locked_keys += f" and {len(workflow_locked_rows) - 5} more"
        raise HTTPException(
            409,
            f"Testcase fields are locked while review or approval is pending ({locked_keys}). "
            "Return the test case for correction before changing its content.",
        )
    if "folder_id" in changes and payload.folder_id is not None:
        if not db.query(models.TestFolder).filter_by(id=payload.folder_id, project_id=project_id).first():
            raise HTTPException(404, "Folder not found in this project")
    if "priority" in changes and payload.priority not in TEST_CASE_PRIORITIES:
        raise HTTPException(400, "Invalid test case priority")
    substantive_changes = changes.intersection({"priority", "test_type", "module_name"})
    for row in rows:
        if "folder_id" in changes:
            row.folder_id = payload.folder_id
        if "tags" in changes:
            row.tag_rows = [models.TestCaseTag(tag=tag) for tag in _normalize_tags(payload.tags or [])]
        if not substantive_changes:
            continue
        draft = row.current_draft_version
        draft_created = False
        if not draft:
            approved = row.current_approved_version
            base_content = {f: getattr(approved, f) if approved else None for f in _CONTENT_FIELDS}
            major, minor = _next_provisional_version_numbers(row)
            draft = models.TestCaseVersion(
                test_case_id=row.id, version_major=major, version_minor=minor, status="Draft",
                author_id=current_user.id, created_at=models.now(),
                source_version_id=approved.id if approved else None, **base_content,
            )
            db.add(draft)
            db.flush()
            draft_created = True
            if approved:
                draft.steps = [models.TestCaseVersionStep(step_no=s.step_no, step_text=s.step_text,
                                                           expected_result=s.expected_result) for s in approved.steps]
            row.current_draft_version_id = draft.id
        content_changed = bool(substantive_changes) and draft.status not in ("In Review", "Recommendation Pending")
        if content_changed:
            if "priority" in changes:
                draft.priority = payload.priority
            if "test_type" in changes:
                draft.test_type = payload.test_type
            if "module_name" in changes:
                draft.module_name = payload.module_name
            db.add(_case_workflow_action(
                row.id, current_user, "Draft updated",
                "Test case classification was changed in a bulk update.",
            ))
        if draft_created or content_changed:
            _sync_case_mirror(row, draft)
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.post("/projects/{project_id}/test-cases/bulk-delete", response_model=List[schemas.TestCaseOut])
def bulk_delete_test_cases(project_id: int, payload: schemas.TestCaseBulkDelete,
                           db: Session = Depends(get_db),
                           current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """2026-08 "Recycle Bin" requirement -- see delete_test_case's own
    docstring for the full background. Soft-deletes every selected row
    (is_deleted=True) instead of a real `db.delete()`; still permanently
    blocked (409) for any ever-Approved/Archived/Rejected row, unchanged."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    rows = _selected_project_cases(db, project_id, payload.ids)
    for row in rows:
        _enforce_checkout_lock(row, current_user)
    ever_decided_ids = {
        row_id for (row_id,) in db.query(models.TestCaseVersion.test_case_id).filter(
            models.TestCaseVersion.test_case_id.in_([r.id for r in rows]),
            models.TestCaseVersion.status.in_(("Approved", "Archived", "Rejected")),
        ).all()
    }
    blocked = [row.test_case_key for row in rows if row.id in ever_decided_ids]
    if blocked:
        preview = ", ".join(blocked[:5])
        suffix = "…" if len(blocked) > 5 else ""
        # 409, same reasoning as the single-case delete above -- the
        # frontend's own bulk-delete selection (see returnRejectSelectedIds'
        # sibling, deletableSelectedIds) is now built to never include an
        # ever-decided test case in the first place, so reaching this is
        # only ever a race (someone else approved/archived/rejected one of
        # these between selection and submission), not an expected path.
        raise HTTPException(
            409,
            f"{len(blocked)} selected test case(s) have an approval/rejection history and cannot be permanently "
            f"deleted -- archive them instead: {preview}{suffix}",
        )
    now = models.now()
    for row in rows:
        row.is_deleted = True
        row.deleted_by_id = current_user.id
        row.deleted_at = now
        db.add(_case_workflow_action(row.id, current_user, "Moved to Recycle Bin",
                                     "Deleted before approval (bulk) -- recoverable from the Recycle Bin until purged."))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.get("/projects/{project_id}/test-cases/recycle-bin", response_model=pagination.Page[schemas.TestCaseListOut])
def list_recycle_bin(project_id: int, params: pagination.PageParams = Depends(),
                     db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """2026-08 "Recycle Bin" requirement -- everything currently soft-
    deleted for this project (delete_test_case/bulk_delete_test_cases),
    awaiting either restore_test_case_from_recycle_bin or a QA Lead Group
    purge. Deliberately a flat list, not folder/status-filtered like the
    main list endpoint -- a case can only ever land here pre-approval, so
    there's no meaningful "queue" to further split it by; search still
    applies so a specific case can be found by key/scenario. Same read
    access as the main list (no extra role gate) -- restore/purge below are
    where the actual authorization differences live."""
    _get_project_or_404(db, project_id)
    q = (
        db.query(models.TestCase)
        .filter(models.TestCase.project_id == project_id, models.TestCase.is_deleted == True)  # noqa: E712 - Oracle requires = 1, not IS 1
        .options(*_LIST_CASE_EAGER_LOADS)
    )
    q = pagination.apply_search(
        q, params,
        models.TestCase.test_case_key, models.TestCase.test_scenario, models.TestCase.epic_id,
        models.TestCase.cr_number, models.TestCase.feature_id, models.TestCase.user_story_id,
        models.TestCase.module_name,
    )
    q = pagination.apply_sort(
        q, params,
        sortable={"test_case_key": models.TestCase.test_case_key, "priority": models.TestCase.priority},
        default_column=models.TestCase.deleted_at, id_column=models.TestCase.id,
    )
    result = pagination.paginate(q, params)
    return pagination.to_page_response(result, params)


@router.post("/test-cases/{case_id}/restore-from-recycle-bin", response_model=schemas.TestCaseOut)
def restore_test_case_from_recycle_bin(case_id: int, db: Session = Depends(get_db),
                                       current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """2026-08 "Recycle Bin" requirement -- undoes a delete_test_case/
    bulk_delete_test_cases soft-delete, same Author-tier access that could
    delete it in the first place (require_can_author_repository) -- not
    restricted to the QA Lead Group the way purging is, since undoing your
    own (or a teammate's) accidental delete isn't a governance decision."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_author_repository(db, obj.project_id, current_user)
    if not obj.is_deleted:
        raise HTTPException(400, "This test case is not in the Recycle Bin")
    obj.is_deleted = False
    obj.deleted_by_id = None
    obj.deleted_at = None
    obj.deleted_reason = None
    db.add(_case_workflow_action(obj.id, current_user, "Restored from Recycle Bin",
                                 "Restored -- no re-approval required, it was never approved."))
    db.commit()
    db.refresh(obj)
    return obj


@router.post("/projects/{project_id}/test-cases/bulk-restore-from-recycle-bin", response_model=List[schemas.TestCaseOut])
def bulk_restore_test_cases_from_recycle_bin(project_id: int, payload: schemas.TestCaseBulkRestoreFromRecycleBin,
                                             db: Session = Depends(get_db),
                                             current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Bulk counterpart to restore_test_case_from_recycle_bin above."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    rows = _selected_project_cases(db, project_id, payload.ids, include_deleted=True)
    not_deleted = [row.test_case_key for row in rows if not row.is_deleted]
    if not_deleted:
        preview = ", ".join(not_deleted[:5])
        suffix = "…" if len(not_deleted) > 5 else ""
        raise HTTPException(400, f"{len(not_deleted)} selected test case(s) are not in the Recycle Bin: {preview}{suffix}")
    for row in rows:
        row.is_deleted = False
        row.deleted_by_id = None
        row.deleted_at = None
        row.deleted_reason = None
        db.add(_case_workflow_action(row.id, current_user, "Restored from Recycle Bin",
                                     "Restored (bulk) -- no re-approval required, it was never approved."))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.delete("/test-cases/{case_id}/purge")
def purge_test_case(case_id: int, db: Session = Depends(get_db),
                    current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """2026-08 "Recycle Bin" requirement -- "only QA lead can clear from
    recycle bin." The only remaining code path that issues a real,
    irreversible `db.delete()` -- everything else in this file that used to
    hard-delete now soft-deletes instead (see delete_test_case's own
    docstring). Requires the case to already be in the Recycle Bin -- a live
    (non-deleted) case can never be purged directly, it must go through
    delete_test_case first."""
    obj = get_or_404(db, models.TestCase, case_id, "Test Case")
    _require_active_project(_get_project_or_404(db, obj.project_id))
    require_can_manage_repository_governance(current_user)
    if not obj.is_deleted:
        raise HTTPException(400, "This test case is not in the Recycle Bin -- delete it first")
    # Logged before the delete -- entity_id on ApprovalAction is a plain
    # (unconstrained) integer column, deliberately so an audit row can
    # outlive the entity it refers to for exactly this kind of terminal
    # action.
    db.add(_case_workflow_action(obj.id, current_user, "Purged from Recycle Bin",
                                 "Permanently deleted by the QA Lead Group -- irreversible."))
    db.delete(obj)
    db.commit()
    return {"ok": True}


@router.post("/projects/{project_id}/test-cases/bulk-purge")
def bulk_purge_test_cases(project_id: int, payload: schemas.TestCaseBulkPurge,
                          db: Session = Depends(get_db),
                          current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Bulk counterpart to purge_test_case above -- also QA Lead Group only,
    also permanent. Powers the Recycle Bin view's "Empty selected"/"Empty
    Recycle Bin" action."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_manage_repository_governance(current_user)
    rows = _selected_project_cases(db, project_id, payload.ids, include_deleted=True)
    not_deleted = [row.test_case_key for row in rows if not row.is_deleted]
    if not_deleted:
        preview = ", ".join(not_deleted[:5])
        suffix = "…" if len(not_deleted) > 5 else ""
        raise HTTPException(400, f"{len(not_deleted)} selected test case(s) are not in the Recycle Bin: {preview}{suffix}")
    purged_ids = [row.id for row in rows]
    for row in rows:
        db.add(_case_workflow_action(row.id, current_user, "Purged from Recycle Bin",
                                     "Permanently deleted by the QA Lead Group (bulk) -- irreversible."))
        db.delete(row)
    db.commit()
    return {"purged": len(purged_ids), "ids": purged_ids}


@router.post("/projects/{project_id}/test-cases/bulk-approve", response_model=List[schemas.TestCaseOut])
def bulk_approve_test_cases(project_id: int, payload: schemas.TestCaseBulkApprove,
                            db: Session = Depends(get_db),
                            current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """QA-Lead-tier bulk FINAL decision (2026-08 Approval Workflow refactor)
    -- approves several "Review Completed" draft versions in one atomic
    decision (always a minor-version approval -- use the single-case review
    endpoint for a major bump, which needs its own per-case justification).
    Reachable by system QA Lead/Admin always, or a QA Engineer holding
    project role Project Lead on this project (OLD-path rows, "Review
    Completed") -- require_can_give_final_approval does that check. NEW-path
    rows ("QA Lead Approval Pending") instead require membership in the QA
    Lead Group (QA_LEAD/CHIEF_MANAGER_QA/AGM_QA), same as the single-case
    review endpoint's NEW-path Stage 2 branch -- see TEST_CASE_NEW_STATUSES
    in constants.py. The whole selection must be homogeneously one path or
    the other (validated below) so a single "who is allowed to do this"
    check always applies to the entire batch, same all-or-nothing convention
    as every other validation this endpoint runs before committing anything.
    See bulk_recommend_test_cases below for the Stage-1 bulk equivalent.
    One approver message is deliberately reused for every case-specific
    audit row, so each testcase retains a complete history without asking
    the lead to repeat the same message."""
    _require_active_project(_get_project_or_404(db, project_id))
    rows = _selected_project_cases(db, project_id, payload.ids)
    _lock_current_drafts(db, rows)
    operation_label = "Bulk approval" if len(rows) > 1 else "Approval"
    comments = payload.comments.strip()
    if not comments:
        raise HTTPException(400, "Enter one approval message for the selected test cases")
    if len(comments) > 5000:
        raise HTTPException(400, "Approval message cannot exceed 5,000 characters")
    not_pending = [row.test_case_key for row in rows if not row.current_draft_version
                   or row.current_draft_version.status not in ("Review Completed", "QA Lead Approval Pending")]
    if not_pending:
        preview = ", ".join(not_pending[:5])
        suffix = "…" if len(not_pending) > 5 else ""
        raise HTTPException(
            400,
            f"{operation_label} stopped because {len(not_pending)} selected test case(s) are not pending "
            f"QA Lead final approval: {preview}{suffix}",
        )
    statuses_present = {row.current_draft_version.status for row in rows}
    if len(statuses_present) > 1:
        raise HTTPException(
            400,
            f"{operation_label} stopped because the selection mixes OLD-workflow (\"Review Completed\") and "
            "NEW-workflow (\"QA Lead Approval Pending\") test cases -- select one group at a time.",
        )
    is_old_path = statuses_present == {"Review Completed"}
    if is_old_path:
        require_can_give_final_approval(db, project_id, current_user)
    elif not current_user.has_role(Role.QA_LEAD, Role.CHIEF_MANAGER_QA, Role.AGM_QA):
        raise HTTPException(403, "Final approval is available only to the QA Lead Group")
    self_authored = [row.test_case_key for row in rows if row.current_draft_version.author_id == current_user.id]
    if self_authored:
        preview = ", ".join(self_authored[:5])
        raise HTTPException(
            403,
            f"GOV-002: you authored the pending draft on {len(self_authored)} selected test case(s) and cannot "
            f"approve your own work: {preview}",
        )
    # NEW-path only (see review_test_case's identical fix for the single-case
    # endpoint, same reported gap) -- also block whoever submitted or
    # Stage-1-recommended a row from being its own Stage 2 approver, not just
    # whoever authored its content. OLD-path GOV-002 stays author-only.
    if not is_old_path:
        self_acted = [row.test_case_key for row in rows
                      if current_user.id in (row.current_draft_version.submitted_by_id,
                                              row.current_draft_version.reviewed_by_id)]
        if self_acted:
            preview = ", ".join(self_acted[:5])
            raise HTTPException(
                403,
                f"GOV-002: you already acted on {len(self_acted)} selected test case(s) at an earlier stage and "
                f"cannot also record their Stage 2 decision: {preview}",
            )
    previous_state = "Review Completed" if is_old_path else "QA Lead Approval Pending"
    for row in rows:
        draft = row.current_draft_version
        draft.status = "Approved"
        draft.qa_lead_decided_by_id = current_user.id
        draft.qa_lead_decided_at = models.now()
        draft.qa_lead_decision_comments = comments
        row.current_approved_version_id = draft.id
        row.current_draft_version_id = None
        _sync_case_mirror(row, draft)
        db.add(_case_workflow_action(row.id, current_user, "Approved & Activated", comments,
                                     previous_state=previous_state, new_state="Approved"))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.post("/projects/{project_id}/test-cases/bulk-recommend", response_model=List[schemas.TestCaseOut])
def bulk_recommend_test_cases(project_id: int, payload: schemas.TestCaseBulkRecommend,
                              db: Session = Depends(get_db),
                              current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Stage-1-tier bulk equivalent of bulk_approve_test_cases above --
    recommends several pending draft versions for QA Lead final approval in
    one atomic decision. OLD-path rows ("In Review") require system QA
    Lead/Admin, or a QA Engineer holding project role Reviewer/Project
    Lead/Owner on this project (require_can_review_repository). NEW-path
    rows ("Recommendation Pending") instead require membership in the QA
    Group (QA_ENGINEER), same as the single-case review endpoint's NEW-path
    Stage 1 branch -- see TEST_CASE_NEW_STATUSES in constants.py. The whole
    selection must be homogeneously one path or the other (validated below),
    same reasoning as bulk_approve_test_cases above. comments is optional
    (APR-004: recommendation comments aren't mandatory, unlike return/reject)."""
    _require_active_project(_get_project_or_404(db, project_id))
    rows = _selected_project_cases(db, project_id, payload.ids)
    _lock_current_drafts(db, rows)
    operation_label = "Bulk recommendation" if len(rows) > 1 else "Recommendation"
    comments = (payload.comments or "").strip()
    if len(comments) > 5000:
        raise HTTPException(400, "Comment cannot exceed 5,000 characters")
    not_pending = [row.test_case_key for row in rows if not row.current_draft_version
                   or row.current_draft_version.status not in ("In Review", "Recommendation Pending")]
    if not_pending:
        preview = ", ".join(not_pending[:5])
        suffix = "…" if len(not_pending) > 5 else ""
        raise HTTPException(
            400,
            f"{operation_label} stopped because {len(not_pending)} selected test case(s) are not pending "
            f"QA recommendation: {preview}{suffix}",
        )
    statuses_present = {row.current_draft_version.status for row in rows}
    if len(statuses_present) > 1:
        raise HTTPException(
            400,
            f"{operation_label} stopped because the selection mixes OLD-workflow (\"In Review\") and "
            "NEW-workflow (\"Recommendation Pending\") test cases -- select one group at a time.",
        )
    is_old_path = statuses_present == {"In Review"}
    if is_old_path:
        require_can_review_repository(db, project_id, current_user)
    elif not current_user.has_role(Role.QA_ENGINEER):
        raise HTTPException(403, "Stage 1 recommendation is available only to the QA Group")
    self_authored = [row.test_case_key for row in rows if row.current_draft_version.author_id == current_user.id]
    if self_authored:
        preview = ", ".join(self_authored[:5])
        raise HTTPException(
            403,
            f"GOV-002: you authored the pending draft on {len(self_authored)} selected test case(s) and cannot "
            f"recommend your own work: {preview}",
        )
    # NEW-path only -- reported directly: a non-author who submitted a draft
    # was then immediately able to recommend the very item they'd just
    # submitted, since this check only ever looked at author_id. OLD-path
    # GOV-002 stays author-only.
    if not is_old_path:
        self_submitted = [row.test_case_key for row in rows
                           if row.current_draft_version.submitted_by_id == current_user.id]
        if self_submitted:
            preview = ", ".join(self_submitted[:5])
            raise HTTPException(
                403,
                f"GOV-002: you submitted the pending draft on {len(self_submitted)} selected test case(s) and "
                f"cannot also record its Stage 1 decision: {preview}",
            )
    management_ids = _stage2_approver_ids(db) if is_old_path else _qa_lead_group_ids(db)
    if not management_ids:
        raise HTTPException(400, "No active QA Lead Group approver is configured")
    previous_state = "In Review" if is_old_path else "Recommendation Pending"
    new_state = "Review Completed" if is_old_path else "QA Lead Approval Pending"
    for row in rows:
        draft = row.current_draft_version
        draft.assigned_qa_lead_id = None
        draft.status = new_state
        draft.reviewed_by_id = current_user.id
        draft.reviewed_at = models.now()
        draft.review_comments = comments or (
            "Recommended for QA Lead final approval (bulk recommend)."
            if len(rows) > 1 else "Recommended for QA Lead final approval."
        )
        _sync_case_mirror(row, draft)
        db.add(_case_workflow_action(row.id, current_user, "Recommended for approval", draft.review_comments,
                                     previous_state=previous_state, new_state=new_state))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


def _bulk_new_path_decision_rows(db: Session, project_id: int, payload_ids: List[int],
                                  current_user: models.User, action_label: str) -> List[models.TestCase]:
    """Shared validation for bulk_return_test_cases/bulk_reject_test_cases
    below -- both are NEW-workflow-only (see their own docstrings), so this
    factors out the parts identical to bulk_recommend_test_cases/
    bulk_approve_test_cases above minus the OLD-path branch: selection must
    be entirely one NEW-path checkpoint ("Recommendation Pending" or "QA
    Lead Approval Pending"), the caller must hold the matching group role,
    and GOV-002 excludes the author plus whoever already acted at an earlier
    stage (submitter for Stage 1, submitter+Stage-1-reviewer for Stage 2)."""
    rows = _selected_project_cases(db, project_id, payload_ids)
    _lock_current_drafts(db, rows)
    operation_label = f"Bulk {action_label}" if len(rows) > 1 else action_label.capitalize()
    _new_pending = ("Recommendation Pending", "QA Lead Approval Pending")
    not_pending = [row.test_case_key for row in rows if not row.current_draft_version
                   or row.current_draft_version.status not in _new_pending]
    if not_pending:
        preview = ", ".join(not_pending[:5])
        suffix = "…" if len(not_pending) > 5 else ""
        raise HTTPException(
            400,
            f"{operation_label} stopped because {len(not_pending)} selected test case(s) are not pending "
            f"a QA Group/QA Lead Group decision: {preview}{suffix}",
        )
    statuses_present = {row.current_draft_version.status for row in rows}
    if len(statuses_present) > 1:
        raise HTTPException(
            400,
            f"{operation_label} stopped because the selection mixes Stage 1 (\"Recommendation Pending\") and "
            f"Stage 2 (\"QA Lead Approval Pending\") test cases -- select one stage at a time.",
        )
    is_stage1 = statuses_present == {"Recommendation Pending"}
    if is_stage1 and not current_user.has_role(Role.QA_ENGINEER):
        raise HTTPException(403, "Stage 1 decisions are available only to the QA Group")
    if not is_stage1 and not current_user.has_role(Role.QA_LEAD, Role.CHIEF_MANAGER_QA, Role.AGM_QA):
        raise HTTPException(403, "Stage 2 decisions are available only to the QA Lead Group")
    self_authored = [row.test_case_key for row in rows if row.current_draft_version.author_id == current_user.id]
    if self_authored:
        preview = ", ".join(self_authored[:5])
        raise HTTPException(
            403,
            f"GOV-002: you authored the pending draft on {len(self_authored)} selected test case(s) and cannot "
            f"{action_label} your own work: {preview}",
        )
    if is_stage1:
        self_acted = [row.test_case_key for row in rows if row.current_draft_version.submitted_by_id == current_user.id]
    else:
        self_acted = [row.test_case_key for row in rows
                      if current_user.id in (row.current_draft_version.submitted_by_id,
                                              row.current_draft_version.reviewed_by_id)]
    if self_acted:
        preview = ", ".join(self_acted[:5])
        raise HTTPException(
            403,
            f"GOV-002: you already acted on {len(self_acted)} selected test case(s) at an earlier stage and "
            f"cannot also {action_label} them: {preview}",
        )
    return rows


@router.post("/projects/{project_id}/test-cases/bulk-return", response_model=List[schemas.TestCaseOut])
def bulk_return_test_cases(project_id: int, payload: schemas.TestCaseBulkReturn,
                            db: Session = Depends(get_db),
                            current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """NEW-workflow-only bulk equivalent of review_test_case's single-case
    RETURN decision -- reported directly, alongside the equivalent bulk
    Reject request below, as the missing bulk counterparts to the existing
    single-case "Recommend Approval / Return for Correction / Reject" panel:
    the first of those three already had bulk-recommend/bulk-approve, these
    two close the gap for the other two decisions. Sends each selected
    row's current draft back to its author ("Returned by QA" from Stage 1,
    "Returned by QA Lead" from Stage 2) with one shared reason. See
    _bulk_new_path_decision_rows above for the shared validation both this
    and bulk_reject_test_cases run first."""
    _require_active_project(_get_project_or_404(db, project_id))
    comments = payload.comments.strip()
    if not comments:
        raise HTTPException(400, "A reason is required when returning test cases for changes")
    if len(comments) > 5000:
        raise HTTPException(400, "Reason cannot exceed 5,000 characters")
    rows = _bulk_new_path_decision_rows(db, project_id, payload.ids, current_user, "return")
    is_stage1 = rows[0].current_draft_version.status == "Recommendation Pending"
    previous_state = "Recommendation Pending" if is_stage1 else "QA Lead Approval Pending"
    new_state = "Returned by QA" if is_stage1 else "Returned by QA Lead"
    for row in rows:
        draft = row.current_draft_version
        draft.status = new_state
        draft.reviewed_by_id = current_user.id
        draft.reviewed_at = models.now()
        draft.review_comments = comments
        if not is_stage1:
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments
        _sync_case_mirror(row, draft)
        db.add(_case_workflow_action(row.id, current_user, "Returned for correction", comments,
                                     previous_state=previous_state, new_state=new_state))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


@router.post("/projects/{project_id}/test-cases/bulk-reject", response_model=List[schemas.TestCaseOut])
def bulk_reject_test_cases(project_id: int, payload: schemas.TestCaseBulkReject,
                            db: Session = Depends(get_db),
                            current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """NEW-workflow-only bulk equivalent of review_test_case's single-case
    REJECT decision (terminal) -- see bulk_return_test_cases immediately
    above for the shared background/reasoning."""
    _require_active_project(_get_project_or_404(db, project_id))
    comments = payload.comments.strip()
    if not comments:
        raise HTTPException(400, "A reason is required when rejecting test cases")
    if len(comments) > 5000:
        raise HTTPException(400, "Reason cannot exceed 5,000 characters")
    rows = _bulk_new_path_decision_rows(db, project_id, payload.ids, current_user, "reject")
    is_stage1 = rows[0].current_draft_version.status == "Recommendation Pending"
    previous_state = "Recommendation Pending" if is_stage1 else "QA Lead Approval Pending"
    for row in rows:
        draft = row.current_draft_version
        draft.status = "Rejected"
        draft.reviewed_by_id = current_user.id
        draft.reviewed_at = models.now()
        draft.review_comments = comments
        if not is_stage1:
            draft.qa_lead_decided_by_id = current_user.id
            draft.qa_lead_decided_at = models.now()
            draft.qa_lead_decision_comments = comments
        _sync_case_mirror(row, draft)
        db.add(_case_workflow_action(row.id, current_user, "Rejected", comments,
                                     previous_state=previous_state, new_state="Rejected"))
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


# ---- xlsx import ----
# Maps the attached template's exact column headers (case/whitespace
# tolerant -- see _normalize_header) to the field each one feeds. Any column
# not in this map is ignored rather than rejected, so the template can gain
# extra columns later without breaking the parser.
_HEADER_MAP = {
    "test case id": "test_case_key",
    "epic id": "epic_id",
    "cr number": "cr_number",
    "cr id": "cr_number",
    "change request": "cr_number",
    "feature id": "feature_id",
    "user story id": "user_story_id",
    "test type": "test_type",
    "module name (if any)": "module_name",
    "test scenario": "test_scenario",
    "pre-condition (if any)": "pre_condition",
    "test case description": "description",
    "step no.": "step_no_label",
    "steps": "step_text",
    "expected result": "expected_result",
    "priority (critical/high/medium/low)": "priority",
    "tags": "tags",
    "labels": "tags",
    "actual result": "actual_result",
    "status (pass/fail/blocked/na/retest passed)": "status",
    "test run artifacts": "test_run_artifacts",
    "defect id (if any)": "defect_id",
}

# Fields captured only from the first row of a test case's block (the
# template leaves these blank on every subsequent step row -- see the
# template's own merged-cell-style layout).
_CASE_LEVEL_FIELDS = [
    "epic_id", "cr_number", "feature_id", "user_story_id", "test_type", "module_name",
    "test_scenario", "pre_condition", "description", "priority", "tags",
]
_EXECUTION_LEVEL_FIELDS = ["actual_result", "status", "test_run_artifacts", "defect_id"]


def _normalize_header(h) -> str:
    return str(h or "").strip().lower()


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


@router.get("/import-template")
def download_import_template(current_user: models.User = Depends(get_current_user)):
    """Serves the canonical "Test Cases - CR-XX - Template" xlsx that
    import_test_cases() below expects, so users import test cases using
    this template only instead of an arbitrary spreadsheet. Any
    authenticated user may download it -- it carries no project data,
    just the required header row and example rows."""
    if not os.path.exists(_IMPORT_TEMPLATE_PATH):
        raise HTTPException(404, "Import template file is missing on the server")
    return FileResponse(
        _IMPORT_TEMPLATE_PATH,
        filename="Test Case Import Template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/projects/{project_id}/import-xlsx", response_model=schemas.TestCaseImportResult)
async def import_test_cases(project_id: int, file: UploadFile = File(...), folder_id: Optional[int] = Form(None),
                             db: Session = Depends(get_db),
                             current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES))):
    """Parses the attached "Test Cases - CR-XX - Template" xlsx format: one
    row per test step, with the test case's own descriptive fields (Epic ID,
    Feature ID, Test Scenario, Priority, etc.) filled in only on the first
    row of each test case's block and left blank on every subsequent step
    row -- a new block starts whenever the "Test Case ID" column is non-empty
    again. Every imported definition enters as a first (1.0) TestCaseVersion
    via the same _create_case_with_first_draft helper create_test_case uses
    (IO-001/IO-002). Every imported testcase remains Draft regardless of
    step completeness; its author must explicitly select Submit for review
    (single or bulk) after checking the imported definition. Uploading data
    must never itself constitute a workflow submission. Execution-result
    columns are deliberately not imported into a cycle: only an Approved
    version may enter Test Execution."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    if folder_id:
        if not db.query(models.TestFolder).filter_by(id=folder_id, project_id=project_id).first():
            raise HTTPException(404, "Folder not found in this project")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(400, "Could not read this file as an Excel (.xlsx) workbook")
    ws = wb.worksheets[0]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(400, "Sheet has no header row")
    col_fields: dict = {}
    for idx, header in enumerate(header_row):
        field = _HEADER_MAP.get(_normalize_header(header))
        if field:
            col_fields[idx] = field

    if "test_case_key" not in col_fields.values():
        raise HTTPException(400, "Could not find a 'Test Case ID' column -- is this the right template?")

    created_test_cases = 0
    imported_executions = 0
    skipped_rows = 0
    errors: List[str] = []
    imported_source_keys: set[str] = set()

    def flush_group(case_row: dict, step_rows: List[dict], source_row: int):
        nonlocal created_test_cases, skipped_rows
        source_key = str(case_row.get("test_case_key") or "").strip()
        if source_key and source_key.casefold() in imported_source_keys:
            group_rows = max(1, len(step_rows))
            end_row = source_row + group_rows - 1
            row_label = f"Rows {source_row}-{end_row}" if end_row > source_row else f"Row {source_row}"
            errors.append(
                f"{row_label}: source Test Case ID '{source_key}' occurs more than once in this workbook; "
                "the duplicate block was skipped."
            )
            skipped_rows += group_rows
            return
        if source_key:
            imported_source_keys.add(source_key.casefold())
        content = {f: case_row.get(f) for f in _CONTENT_FIELDS}
        steps = [
            {"step_no": i + 1, "step_text": s.get("step_text"), "expected_result": s.get("expected_result")}
            for i, s in enumerate(step_rows)
            if s.get("step_text") or s.get("expected_result")
        ]
        tags = _normalize_tags((case_row.get("tags") or "").split(","))
        tc = _create_case_with_first_draft(db, project_id, folder_id, content, tags, steps, current_user)
        source_note = f"Imported from Excel row {source_row}" + (f" (source Test Case ID: {source_key})" if source_key else "")
        db.add(_case_workflow_action(
            tc.id, current_user, "Imported as Draft",
            f"{source_note}; saved as Draft. The author must explicitly submit it for Reviewer recommendation.",
            previous_state=None, new_state="Draft",
        ))
        created_test_cases += 1

    current_case: Optional[dict] = None
    current_steps: List[dict] = []
    current_case_source_row = 0
    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None for c in row):
            continue
        parsed = {}
        for idx, field in col_fields.items():
            if idx < len(row):
                parsed[field] = _clean(row[idx])

        starts_new_case = bool(parsed.get("test_case_key"))
        if starts_new_case:
            if current_case is not None:
                flush_group(current_case, current_steps, current_case_source_row)
            current_case = {f: parsed.get(f) for f in _CASE_LEVEL_FIELDS + _EXECUTION_LEVEL_FIELDS}
            current_case["test_case_key"] = parsed.get("test_case_key")
            current_steps = []
            current_case_source_row = excel_row_number

        # The supplied template deliberately allows Test Case ID to be left
        # blank (the example block has six populated steps but no ID). When
        # the first populated row clearly contains case-level information,
        # treat it as a valid new test case and generate the ID in
        # flush_group instead of reporting every following step as orphaned.
        if current_case is None and any(parsed.get(f) for f in _CASE_LEVEL_FIELDS):
            current_case = {f: parsed.get(f) for f in _CASE_LEVEL_FIELDS + _EXECUTION_LEVEL_FIELDS}
            current_case["test_case_key"] = None
            current_steps = []
            current_case_source_row = excel_row_number

        if current_case is None:
            # A step-only row appeared before any Test Case ID was ever seen
            # -- malformed sheet, nothing to attach it to.
            skipped_rows += 1
            step_label = parsed.get("step_no_label") or "unlabelled step"
            errors.append(
                f"Row {excel_row_number}: {step_label} was skipped because it appears before any Test Case ID "
                "or populated test-case details. Fill in Test Case ID, Epic ID, Test Scenario, or another case field."
            )
            continue
        if parsed.get("step_text") or parsed.get("expected_result"):
            current_steps.append({"step_text": parsed.get("step_text"), "expected_result": parsed.get("expected_result")})

    if current_case is not None:
        flush_group(current_case, current_steps, current_case_source_row)

    if created_test_cases == 0 and skipped_rows == 0:
        errors.append("No test cases were found. Confirm that data starts below the header row and uses the standard template columns.")

    db.commit()
    failure_reason = None
    if created_test_cases == 0:
        failure_reason = errors[0] if errors else (
            "No test cases were created. The workbook did not contain a recognizable test-case block."
        )

    return schemas.TestCaseImportResult(
        created_test_cases=created_test_cases, imported_executions=imported_executions,
        skipped_rows=skipped_rows, errors=errors, failure_reason=failure_reason,
    )


@router.post("/projects/{project_id}/import-xlsx/jobs")
async def queue_test_case_import(
    project_id: int,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    folder_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_roles(*_AUTHOR_ROLES)),
):
    """Upload immediately, then parse and insert outside the HTTP request."""
    _require_active_project(_get_project_or_404(db, project_id))
    require_can_author_repository(db, project_id, current_user)
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "The uploaded Excel file is empty")
    filename = os.path.basename(file.filename or "testcases.xlsx")
    user_id = current_user.id

    def process(job_id: str):
        with SessionLocal() as worker_db:
            worker_user = worker_db.query(models.User).get(user_id)
            if not worker_user:
                raise RuntimeError("The user who started this import no longer exists")
            jobs.update(job_id, progress=15)
            upload = UploadFile(filename=filename, file=io.BytesIO(raw))
            result = asyncio.run(import_test_cases(
                project_id,
                upload,
                folder_id,
                worker_db,
                worker_user,
            ))
            return result.model_dump(mode="json")

    return jobs.enqueue(background_tasks, "TESTCASE_XLSX_IMPORT", user_id, process)
